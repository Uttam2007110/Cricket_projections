# -*- coding: utf-8 -*-
"""
Created on Tue Apr 18 16:46:51 2023
Generating fantasy xPts for dream 11 based on projections
@author: Subramanya.Ganti
"""
#%% choose the teams which are playing
import pandas as pd
import numpy as np
from scipy.stats import truncnorm
from pulp import LpMaximize, LpProblem, lpSum, LpVariable, GLPK
from itertools import chain
import datetime as dt
from openpyxl import load_workbook
import random
from collections import Counter
pd.options.mode.chained_assignment = None  # default='warn'
np.seterr(all='ignore')

comp = 'tests'; year = '24'; unique_combos = 11
#if another league data is used as a proxy then set 1
home=[]; opps=[]; venue = []; team_bat_first = []; proxy = 0; custom = 0
#date based game selection if 0, else specific gameweek or entire season
gw = 0; write = 0
#select teams manually
home = ['Australia']; opps = ['India']; venue = ["Adelaide Oval"]; team_bat_first = ['']
#home = ['Zimbabwe']; opps = ['Pakistan']; venue = ["Queens Sports Club"]; team_bat_first = ['']
#select custom date
#custom = dt.datetime(2024,11,6) #year,month,date
#type of scoring system, default dream 11
coversoff = 0; ex22 = 0; cricdraft = 0
#the length of the reduced game due to rain or other factors, full game is 1
reduced_length = 1

#frauds like ben stokes who bowl whenever they feel like it
not_bowling_list = ['H Klaasen','SIR Dunkley','GM Harris','D Ferreira','RK Singh','RR Hendricks','Tilak Varma','T Stubbs','SA Yadav',
                    'Abhishek Sharma','TH David','WG Jacks','N Pooran','JG Bethell','SE Rutherford','MNK Fernando','C Webb','SW Bates',
                    'JI Rodrigues','R McKenna','HJ Armitage']
#frauds who suddenly decide to bat in a different position
custom_position_list = [['KL Rahul',6],['NA McSweeney',2],['Nithish Kumar Reddy',8],['SPD Smith',4],['NM Lyon',10],['V Kohli',4]]
#designated wicketkeeper in a game
designated_keeper_list = ['AT Carey','RR Pant','T Marumani','Mohammad Rizwan']

#%% find projections for the games in question
from usage import *
path = 'C:/Users/Subramanya.Ganti/Downloads/cricket'
if(proxy == 1): input_file = f"{path}/projections/{comp}_proxy_projections.xlsx"
else: input_file = f"{path}/projections/{comp}_projections.xlsx"
input_file1 = f"{path}/summary/{comp}_summary.xlsx"
output_dump = f"{path}/outputs/{comp}_{year}_game.xlsx"
game_avg = []
home_win = []

custom_position_list = pd.DataFrame(custom_position_list, columns=['player', 'position'])
custom_position_list = custom_position_list.sort_values(by=['position'],ascending=True)
with pd.ExcelWriter(f'{path}/custom_positions.xlsx') as writer:
    custom_position_list.to_excel(writer, sheet_name="custom positions", index=False)

if(comp=='hundred' or comp=='hundredw'):
    f = (5/6); #hundred
elif(comp=='odi' or comp=='odiw' or comp=='odiq' or comp=='rlc' or comp=='rhf'):
    f = 2.5;   #odi
elif(comp=='tests' or comp == 'cc' or comp == 'shield' or comp == 'testsw'):
    f = 11.25; #test
else:
    f = 1;     #assume its a t20 by default
    
def fixtures_file(now,comp,gw):
    fixtures = f"{path}/schedule.xlsx"
    if(comp == 'hundredw'): fixtures = pd.read_excel(fixtures,f'hundred {year}')
    else: fixtures = pd.read_excel(fixtures,f'{comp} {year}')
    temp = dt.datetime(1899, 12, 30)    # Note, not 31st Dec but 30th!
    delta = now - temp
    now = float(delta.days)
    home = []; opps = []

    for x in fixtures.values:
        if(gw == 0 and x[1] == now):
            home.append(x[3])
            opps.append(x[4])
            venue.append(x[5])
        elif(gw == x[2]):
            home.append(x[3])
            opps.append(x[4])
            venue.append(x[5])
            
    return (home,opps)

if(home==[] and opps==[]):
    if(custom !=0): now = custom
    else: now = dt.datetime.now()
    (home,opps) = fixtures_file(now,comp,gw)
else:
    now = dt.datetime.now()

def adj(x,f,bat):
    if(bat == 1):
        #print(x)
        x[6] = x[6]*f
        x[5] = x[5]*f
    else:
        x[4] = x[4]*f
        x[11] = x[11]*f
        
    x[14] = x[14]*f
    x[15] = x[15]*f
    x[16] = x[16]*f
    x[17] = x[17]*f
    x[13] = x[13]*f
    x[12] = 1 - (x[14]+x[15]+x[16]+x[17]+x[13])
    return x

def dismissal_proportions():
    columns = ['wicket type', 'tests', 'odi', 't20']
    data = [('bowled', 0.176193428394296, 0.189325544344132, 0.185424873382105), 
            ('lbw', 0.160074395536268, 0.110993096123208, 0.0766272744325642), 
            ('run out', 0.0255424674519529, 0.0580191184280404, 0.0633089476646033),
            ('stumped', 0.0177309361438314, 0.0220392989909719, 0.0236353404614519),
            ('caught wk', 0.199556760696853, 0.147684189093596, 0.10764762015833),
            ('caught fielder', 0.418174175446978, 0.46888511520805, 0.540073257725744)]
    data = pd.DataFrame(data, columns=columns)
    return data

def get_truncated_normal(mean=0, sd=1, low=0, upp=10):
    return truncnorm((low - mean) / sd, (upp - mean) / sd, loc=mean, scale=sd)

def adj_bowl_usage(t1,t2,bowl_game,bias_spin,bias_pace,factor):
    #print(bowl_game.loc[bowl_game['team']==t1,'usage'].sum())
    #print(bowl_game.loc[bowl_game['team']==t2,'usage'].sum())
    t1_bowl = bowl_game.loc[bowl_game['team']==t1,'usage'].sum()
    t1_spin = bowl_game.loc[(bowl_game['team']==t1)&(bowl_game['bowlType_main']=='spin'),'usage'].sum()
    t1_pace = bowl_game.loc[(bowl_game['team']==t1)&(bowl_game['bowlType_main']=='pace'),'usage'].sum()
    #print(t1_pace,t1_spin)
    t2_bowl = bowl_game.loc[bowl_game['team']==t2,'usage'].sum()
    t2_spin = bowl_game.loc[(bowl_game['team']==t2)&(bowl_game['bowlType_main']=='spin'),'usage'].sum()
    t2_pace = bowl_game.loc[(bowl_game['team']==t2)&(bowl_game['bowlType_main']=='pace'),'usage'].sum()
    
    bowl_game.loc[(bowl_game['team']==t1)&(bowl_game['bowlType_main']=='spin'),'usage'] = (bias_spin/0.37)*bowl_game.loc[(bowl_game['team']==t1)&(bowl_game['bowlType_main']=='spin'),'usage']
    bowl_game.loc[(bowl_game['team']==t1)&(bowl_game['bowlType_main']=='pace'),'usage'] = (bias_pace/0.63)*bowl_game.loc[(bowl_game['team']==t1)&(bowl_game['bowlType_main']=='pace'),'usage']
    bowl_game.loc[(bowl_game['team']==t2)&(bowl_game['bowlType_main']=='spin'),'usage'] = (bias_spin/0.37)*bowl_game.loc[(bowl_game['team']==t2)&(bowl_game['bowlType_main']=='spin'),'usage']
    bowl_game.loc[(bowl_game['team']==t2)&(bowl_game['bowlType_main']=='pace'),'usage'] = (bias_pace/0.63)*bowl_game.loc[(bowl_game['team']==t2)&(bowl_game['bowlType_main']=='pace'),'usage']
    
    delta = 1
    while(delta > 0.00000001):
        if(factor==11.25): bowl_game['usage'] = np.minimum(bowl_game['usage'],0.5)
        else: bowl_game['usage'] = np.minimum(bowl_game['usage'],0.2)
        delta = t1_bowl - bowl_game.loc[bowl_game['team']==t1,'usage'].sum()
        count_t1 = len(bowl_game.loc[bowl_game['team']==t1,'team'])
        bowl_game.loc[bowl_game['team']==t1,'usage'] = bowl_game.loc[bowl_game['team']==t1,'usage'] + (delta/count_t1)
        #print(delta)
    delta = 1
    while(delta > 0.00000001):
        if(factor==11.25): bowl_game['usage'] = np.minimum(bowl_game['usage'],0.5)
        else: bowl_game['usage'] = np.minimum(bowl_game['usage'],0.2)
        delta = t2_bowl - bowl_game.loc[bowl_game['team']==t2,'usage'].sum()
        count_t2 = len(bowl_game.loc[bowl_game['team']==t2,'team'])
        bowl_game.loc[bowl_game['team']==t2,'usage'] = bowl_game.loc[bowl_game['team']==t2,'usage'] + (delta/count_t2)
        #print(delta)        
    return bowl_game

def toss_effects(batting,bowling,bat_first,t1,t2,factor,reduced_length):
    
    if(bat_first==t1): 
        bat_first_runs = 120*reduced_length*factor*(bowling.loc[bowling['team']==t2,'usage']*bowling.loc[bowling['team']==t2,'runs/ball']).sum()
        bat_second_runs = 120*reduced_length*factor*(bowling.loc[bowling['team']==t1,'usage']*bowling.loc[bowling['team']==t1,'runs/ball']).sum()
    elif(bat_first==t2): 
        bat_first_runs = 120*reduced_length*factor*(bowling.loc[bowling['team']==t1,'usage']*bowling.loc[bowling['team']==t1,'runs/ball']).sum()
        bat_second_runs = 120*reduced_length*factor*(bowling.loc[bowling['team']==t2,'usage']*bowling.loc[bowling['team']==t2,'runs/ball']).sum()
    
    #print(bat_first_runs,bat_second_runs)
    
    if(bat_first_runs < bat_second_runs and bat_first == t1): 
        decay = bat_first_runs/bat_second_runs
        #new_bat_usage = decay * (batting.loc[batting['team']==bat_first,'usage'].sum())
        new_bat_usage = decay * (batting.loc[batting['team']==t2,'usage'].sum())
    elif(bat_first_runs < bat_second_runs and bat_first == t2): 
        decay = bat_first_runs/bat_second_runs
        new_bat_usage = decay * (batting.loc[batting['team']==t1,'usage'].sum())
    else: 
        decay = 1
        new_bat_usage = batting.loc[batting['team']==bat_first,'usage'].sum()
    #print(decay,new_bat_usage)
    
    exp = 1.0
    if(bat_first==t2 and decay<1):
        bowling.loc[bowling['team']==t2,'usage'] = bowling.loc[bowling['team']==t2,'usage'] * decay
        while((batting.loc[batting['team']==t1,'usage']*pow(exp,-batting.loc[batting['team']==t1,'batting_order'])).sum() > new_bat_usage):
            #print(exp,(batting.loc[batting['team']==t1,'usage']*batting.loc[batting['team']==t1,'runs/ball']).sum())
            exp += 0.001
        batting.loc[batting['team']==t1,'usage'] = batting.loc[batting['team']==t1,'usage']*pow(exp,-batting.loc[batting['team']==t1,'batting_order'])
        
    elif(bat_first==t1 and decay<1):
        bowling.loc[bowling['team']==t1,'usage'] = bowling.loc[bowling['team']==t1,'usage'] * decay
        while((batting.loc[batting['team']==t2,'usage']*pow(exp,-batting.loc[batting['team']==t2,'batting_order'])).sum() > new_bat_usage):
            exp += 0.001
        batting.loc[batting['team']==t2,'usage'] = batting.loc[batting['team']==t2,'usage']*pow(exp,-batting.loc[batting['team']==t2,'batting_order'])

    return batting,bowling

def rain_effects(reduced_length,factor):
    balls = (10/reduced_length)/(120*factor)
    balls0 = 10/(120*factor)
    wb = 0.08*np.power(0.992478843,reduced_length*120*factor)
    wb0 = 0.08*np.power(0.992478843,120*factor)
    wbf = wb/wb0
    rb = np.log(balls/0.006694)/np.log(5.120267)
    rb0 = np.log(balls0/0.006694)/np.log(5.120267)
    rbf = rb/rb0
    #print(wb,rb,wbf,rbf)
    return wbf,rbf

def base_calculations(a,b,input_file1,input_file,factor,v,tbf,reduced_length):
    t1 = a
    t2 = b
    print(t1,"vs",t2)
    print('Venue -',v)
    if(reduced_length != 1): print("reduced game,",20*factor*reduced_length,"overs per team")
    if(tbf == t1 or tbf == t2): print(tbf,'are batting first')
    else: print('No information regarding the toss')
    venues = pd.read_excel(input_file,'Venue factors')
    
    #square root because the factor is for the game as a whole?
    vrf = (venues.loc[(venues['venue']==v)&(venues['bowl_type']=='all'),'runs'].sum())**0.5
    vrf_pace = (venues.loc[(venues['venue']==v)&(venues['bowl_type']=='pace'),'runs'].sum())**0.5
    vrf_spin = (venues.loc[(venues['venue']==v)&(venues['bowl_type']=='spin'),'runs'].sum())**0.5
    vwf = (venues.loc[(venues['venue']==v)&(venues['bowl_type']=='all'),'wkts'].sum())**0.5
    vwf_pace = (venues.loc[(venues['venue']==v)&(venues['bowl_type']=='pace'),'wkts'].sum())**0.5
    vwf_spin = (venues.loc[(venues['venue']==v)&(venues['bowl_type']=='spin'),'wkts'].sum())**0.5
    bias_spin = (venues.loc[(venues['venue']==v)&(venues['bowl_type']=='spin'),'bias'].sum())
    bias_pace = (venues.loc[(venues['venue']==v)&(venues['bowl_type']=='pace'),'bias'].sum())
    if(vrf==0):vrf=1; vrf_pace=1; vrf_spin=1; bias_spin=0.37; bias_pace=0.63
    if(vwf==0):vwf=1; vwf_pace=1; vwf_spin=1; bias_spin=0.37; bias_pace=0.63
    
    (summary,bat,bowl) = h2h_alt(input_file1,input_file,1,factor)
    rain_wkts,rain_runs = rain_effects(reduced_length,factor)
    
    people = pd.read_excel(f"{path}/people.xlsx",'people')
    handedness_bowl = people[['unique_name','bowlType']]
    bowl = bowl.merge(handedness_bowl, left_on='bowler', right_on='unique_name')
    bowl = bowl.drop(columns=['unique_name'])
    bowl['bowlType'] = bowl['bowlType'].fillna('Right-arm medium')
    bowl['bowlType_main'] = np.where((bowl['bowlType']=='Right-arm offbreak')|(bowl['bowlType']=='Legbreak googly')|(bowl['bowlType']=='Legbreak')|(bowl['bowlType']=='Left-arm wrist-spin')|(bowl['bowlType']=='Slow left-arm orthodox'),'spin', 'pace')
    
    #bowl = adj_bowl_usage(t1,t2,bowl,bias_spin,bias_pace)
    summary = summary.apply(pd.to_numeric, errors='ignore')
    league_avg = (reduced_length*rain_runs)*(summary.loc[(summary['Team']!="Free Agent"),'runs bat'].mean() + summary.loc[(summary['Team']!="Free Agent"),'runs bowl'].mean())/2
    bat['runs/ball'] = bat['runs/ball']*vrf*rain_runs
    bat['SR'] = bat['runs/ball']*100*rain_runs
    bat['xSR'] = bat['xSR']*vrf*rain_runs
    
    #bowl['runs/ball'] = bowl['runs/ball']*vrf
    #bowl['xECON'] = bowl['xECON']*vrf  
    bowl.loc[bowl['bowlType_main']=='spin','runs/ball'] = bowl.loc[bowl['bowlType_main']=='spin','runs/ball']*vrf_spin*rain_runs
    bowl.loc[bowl['bowlType_main']=='spin','xECON'] = bowl.loc[bowl['bowlType_main']=='spin','xECON']*vrf_spin*rain_runs
    bowl.loc[bowl['bowlType_main']=='pace','runs/ball'] = bowl.loc[bowl['bowlType_main']=='pace','runs/ball']*vrf_pace*rain_runs
    bowl.loc[bowl['bowlType_main']=='pace','xECON'] = bowl.loc[bowl['bowlType_main']=='pace','xECON']*vrf_pace*rain_runs
    bowl = bowl[~bowl['bowler'].isin(not_bowling_list)]
    
    bat['wickets/ball'] = bat['wickets/ball']*vwf*rain_wkts
    #bowl['wickets/ball'] = bowl['wickets/ball']*vwf
    bowl.loc[bowl['bowlType_main']=='spin','wickets/ball'] = bowl.loc[bowl['bowlType_main']=='spin','wickets/ball']*vwf_spin*rain_wkts
    bowl.loc[bowl['bowlType_main']=='pace','wickets/ball'] = bowl.loc[bowl['bowlType_main']=='pace','wickets/ball']*vwf_pace*rain_wkts
    
    if(factor<0.75):
        #old logic for balls faced
        ut1 = (bat.loc[bat['team']==t2,'usage'].sum())/(bowl.loc[bowl['team']==t1,'usage'].sum())
        ut2 = (bat.loc[bat['team']==t1,'usage'].sum())/(bowl.loc[bowl['team']==t2,'usage'].sum())
        bowl.loc[bowl['team']==t1,'usage'] = bowl.loc[bowl['team']==t1,'usage'] * ut1
        bowl.loc[bowl['team']==t2,'usage'] = bowl.loc[bowl['team']==t2,'usage'] * ut2
    else:
        #print(t1,"bowling",bowl.loc[bowl['team']==t1,'usage'].sum()*120*factor)
        #print(t2,"bowling",bowl.loc[bowl['team']==t2,'usage'].sum()*120*factor)
        #print(t1,"batting",bat.loc[bat['team']==t1,'usage'].sum()*120*factor)
        #print(t2,"batting",bat.loc[bat['team']==t2,'usage'].sum()*120*factor)
        #print(t1,"bowling wickets",(bowl.loc[bowl['team']==t1,'usage']*bowl.loc[bowl['team']==t1,'wickets/ball']).sum()*120*factor)
        #print(t2,"bowling wickets",(bowl.loc[bowl['team']==t2,'usage']*bowl.loc[bowl['team']==t2,'wickets/ball']).sum()*120*factor)
        #print(t1,"batting wickets",(bat.loc[bat['team']==t1,'usage']*bat.loc[bat['team']==t1,'wickets/ball']).sum()*120*factor)
        #print(t2,"batting wickets",(bat.loc[bat['team']==t2,'usage']*bat.loc[bat['team']==t2,'wickets/ball']).sum()*120*factor)
        
        #new logic for balls faced
        bb_t1 = bowl.loc[bowl['team']==t1,'usage'].sum()
        bb_t2 = bowl.loc[bowl['team']==t2,'usage'].sum()
        bf_t1 = bat.loc[bat['team']==t1,'usage'].sum()
        bf_t2 = bat.loc[bat['team']==t2,'usage'].sum()
        wbo_t1 = (bowl.loc[bowl['team']==t1,'usage']*bowl.loc[bowl['team']==t1,'wickets/ball']).sum()
        wbo_t2 = (bowl.loc[bowl['team']==t2,'usage']*bowl.loc[bowl['team']==t2,'wickets/ball']).sum()
        wba_t1 = (bat.loc[bat['team']==t1,'usage']*bat.loc[bat['team']==t1,'wickets/ball']).sum()
        wba_t2 = (bat.loc[bat['team']==t2,'usage']*bat.loc[bat['team']==t2,'wickets/ball']).sum()
        
        ut3 = min(pow(bb_t1*bf_t2*wba_t2/wbo_t1,0.5),1)
        ut4 = min(pow(bb_t2*bf_t1*wba_t1/wbo_t2,0.5),1)
        ut1b = ut3 / (bowl.loc[bowl['team']==t1,'usage'].sum())
        ut2b = ut4 / (bowl.loc[bowl['team']==t2,'usage'].sum())
        ut1bat = ut4 / (bat.loc[bat['team']==t1,'usage'].sum())
        ut2bat = ut3 / (bat.loc[bat['team']==t2,'usage'].sum())
        bowl.loc[bowl['team']==t1,'usage'] = bowl.loc[bowl['team']==t1,'usage'] * ut1b
        bowl.loc[bowl['team']==t2,'usage'] = bowl.loc[bowl['team']==t2,'usage'] * ut2b
        bat.loc[bat['team']==t2,'usage'] = bat.loc[bat['team']==t2,'usage'] * ut2bat
        bat.loc[bat['team']==t1,'usage'] = bat.loc[bat['team']==t1,'usage'] * ut1bat
        #print(bowl.loc[bowl['team']==t1,'usage'].sum(),bowl.loc[bowl['team']==t2,'usage'].sum(),ut3,ut4)
    
    bowl = adj_bowl_usage(t1,t2,bowl,bias_spin,bias_pace,factor)
    bowl_t2 = (bowl.loc[bowl['team']==t2,'usage']*bowl.loc[bowl['team']==t2,'runs/ball']*factor*120*reduced_length).sum()
    bowl_t1 = (bowl.loc[bowl['team']==t1,'usage']*bowl.loc[bowl['team']==t1,'runs/ball']*factor*120*reduced_length).sum()
    bat_t2 = (bat.loc[bat['team']==t2,'usage']*bat.loc[bat['team']==t2,'runs/ball']*factor*120*reduced_length).sum() + (bowl.loc[bowl['team']==t1,'usage']*bowl.loc[bowl['team']==t1,'extras/ball']*factor*120*reduced_length).sum()
    bat_t1 = (bat.loc[bat['team']==t1,'usage']*bat.loc[bat['team']==t1,'runs/ball']*factor*120*reduced_length).sum() + (bowl.loc[bowl['team']==t2,'usage']*bowl.loc[bowl['team']==t2,'extras/ball']*factor*120*reduced_length).sum()
    #print("bowl t2",bowl_t2);
    #print("bowl t1",bowl_t1)
    #print("bat t2",bat_t2)
    #print("bat t1",bat_t1)
    #print(league_avg)
    s1 = bowl_t2/(league_avg*bowl.loc[bowl['team']==t2,'usage'].sum())
    c1 = bat_t2/(league_avg*bat.loc[bat['team']==t2,'usage'].sum())
    s2 = bowl_t1/(league_avg*bowl.loc[bowl['team']==t1,'usage'].sum())
    c2 = bat_t1/(league_avg*bat.loc[bat['team']==t1,'usage'].sum())
    
    w1 = 0; w2 = 0;
    w1_bowl = 0; w2_bowl = 0;

    #bat = pd.read_excel(input_file,'MDist bat')
    bat = bat.drop(['RSAA'],axis=1)
    bat = bat.drop(['OAA'],axis=1)
    bat = bat.drop(['xballs/wkt'],axis=1)
    #bat = bat.drop(['xSR'],axis=1)
    
    #bowl = pd.read_excel(input_file,'MDist bowl')
    bowl = bowl.drop(['RCAA'],axis=1)
    bowl = bowl.drop(['WTAA'],axis=1)
    #bowl = bowl.drop(['xECON'],axis=1)
    bowl = bowl.drop(['xSR'],axis=1)
    
    bat_game = [["batsman","season","team","usage","balls/wkt","SR","runs/ball","wickets/ball","pp usage","mid usage","setup usage","death usage","dots/ball","1s/ball","2s/ball","3s/ball","4s/ball","6s/ball","xPts","xSR","batting_order"]]
    bowl_game = [["bowler","season","team","usage","ECON","SR","wickets/ball","pp usage","mid usage","setup usage","death usage","runs/ball","dots/ball","1s/ball","2s/ball","3s/ball","4s/ball","6s/ball","extras/ball","xPts","xECON","description","type"]]

    w_avg_bat = bat.loc[(bat['team']!="Free Agent"),'usage']*bat.loc[(bat['team']!="Free Agent"),'wickets/ball']
    w_avg_bowl = bowl.loc[(bowl['team']!="Free Agent"),'usage']*bowl.loc[(bowl['team']!="Free Agent"),'wickets/ball']*(bat.loc[(bat['team']!="Free Agent"),'usage'].sum()/bowl.loc[(bowl['team']!="Free Agent"),'usage'].sum())
    w_avg_bat = w_avg_bat.sum()*120*reduced_length*factor/(summary.shape[0]-1)
    w_avg_bowl = w_avg_bowl.sum()*120*reduced_length*factor/(summary.shape[0]-1)
    w_avg = (w_avg_bat + w_avg_bowl)/2
    #print(w_avg_bat,w_avg_bowl,w_avg)
    #print(bat.columns)
    #print(bowl.columns)

    for x in bat.values:
        if(x[2] == t1):
            x = adj(x,s1,1)
            bat_game.append(x.tolist())
            w1 = w1 + x[3]*x[7]*120*factor*reduced_length
        if(x[2] == t2):
            x = adj(x,s2,1)
            bat_game.append(x.tolist())
            w2 = w2 + x[3]*x[7]*120*factor*reduced_length

    for x in bowl.values:
        if(x[2] == t1):
            x = adj(x,c1,0)
            bowl_game.append(x.tolist())
            w1_bowl = w1_bowl + x[3]*x[6]*120*factor*reduced_length
        if(x[2] == t2):
            x = adj(x,c2,0)
            bowl_game.append(x.tolist())
            w2_bowl = w2_bowl + x[3]*x[6]*120*factor*reduced_length

    bat_game = pd.DataFrame(bat_game)
    bat_game.columns = bat_game.iloc[0];bat_game = bat_game.drop(0)    
    bowl_game = pd.DataFrame(bowl_game)
    bowl_game.columns = bowl_game.iloc[0];bowl_game = bowl_game.drop(0)
    bat_game = bat_game.apply(pd.to_numeric, errors='ignore')
    bowl_game = bowl_game.apply(pd.to_numeric, errors='ignore')
    
    global game_avg,t_runs_t1,t_runs_t2
    t_runs_t1 = (bowl_game.loc[bowl_game['team'] == t2, 'runs/ball'] * bowl_game.loc[bowl_game['team'] == t2, 'usage']).sum() * 120 * factor
    t_runs_t2 = (bowl_game.loc[bowl_game['team'] == t1, 'runs/ball'] * bowl_game.loc[bowl_game['team'] == t1, 'usage']).sum() * 120 * factor
    game_avg.append(round((t_runs_t1+t_runs_t2)/2,2))
    
    bat_game["xSR"] = bat_game["xPts"]
    bowl_game["xECON"] = bowl_game["xPts"]
    #bat_game['wickets/ball'] = bat_game['wickets/ball']*vwf
    #bowl_game['wickets/ball'] = bowl_game['wickets/ball']*vwf
    #hundred has only 5 balls per over
    if(factor==(5/6)): bowl_game['ECON'] = 5*bowl_game['runs/ball']; bowl_game['xECON'] = (5/6)*bowl_game['xECON']
    else: bowl_game['ECON'] = 6*bowl_game['runs/ball']
    
    #print(w1,w2,w1_bowl,w2_bowl,ut3,ut4,w_avg)
    dismissals = dismissal_proportions()
    if(factor == 11.25):
        avg_runouts = w_avg*dismissals.loc[dismissals['wicket type'] == 'run out','tests'].values[0]
    elif(factor == 2.5):
        avg_runouts = w_avg*dismissals.loc[dismissals['wicket type'] == 'run out','odi'].values[0]
    else:
        avg_runouts = w_avg*dismissals.loc[dismissals['wicket type'] == 'run out','t20'].values[0]
      
    #print(avg_runouts)    
    if(factor == 11.25):
        bat_game.loc[bat_game['team'] == t1, 'wickets/ball'] = bat_game.loc[bat_game['team'] == t1, 'wickets/ball']*(w2_bowl+avg_runouts*reduced_length)/w_avg
        bat_game.loc[bat_game['team'] == t2, 'wickets/ball'] = bat_game.loc[bat_game['team'] == t2, 'wickets/ball']*(w1_bowl+avg_runouts*reduced_length)/w_avg
        bat_game['balls/wkt'] = 1/bat_game['wickets/ball']
        bowl_game.loc[bowl_game['team'] == t1, 'wickets/ball'] = bowl_game.loc[bowl_game['team'] == t1, 'wickets/ball']*(w2-avg_runouts*reduced_length)/w_avg
        bowl_game.loc[bowl_game['team'] == t2, 'wickets/ball'] = bowl_game.loc[bowl_game['team'] == t2, 'wickets/ball']*(w1-avg_runouts*reduced_length)/w_avg
        #proj_wkts = (bowl_game['usage']*bowl_game['wickets/ball']*f*120*reduced_length).sum()
    else:
        #number of wickets that would fall in the full quota of overs, not applicable in tests because theres no upper limit for how long you can bat
        w1_bowl = min(w1_bowl/ut3,10)
        w2 = min(w2/ut3,10)
        w2_bowl = min(w2_bowl/ut4,10)
        w1 = min(w1/ut4,10)
        bat_game.loc[bat_game['team'] == t1, 'wickets/ball'] = bat_game.loc[bat_game['team'] == t1, 'wickets/ball']*(w2_bowl+0.35*reduced_length)/w_avg
        bat_game.loc[bat_game['team'] == t2, 'wickets/ball'] = bat_game.loc[bat_game['team'] == t2, 'wickets/ball']*(w1_bowl+0.35*reduced_length)/w_avg
        bat_game['balls/wkt'] = 1/bat_game['wickets/ball']
        bowl_game.loc[bowl_game['team'] == t1, 'wickets/ball'] = bowl_game.loc[bowl_game['team'] == t1, 'wickets/ball']*(w2-0.35*reduced_length)/w_avg
        bowl_game.loc[bowl_game['team'] == t2, 'wickets/ball'] = bowl_game.loc[bowl_game['team'] == t2, 'wickets/ball']*(w1-0.35*reduced_length)/w_avg
        
        #toss function called only if its known who has won the toss
        if(tbf == t1 or tbf == t2):  
            bat_game,bowl_game = toss_effects(bat_game.copy(),bowl_game.copy(),tbf,t1,t2,factor,reduced_length)
        else:
            bat_game_1,bowl_game_1 = toss_effects(bat_game.copy(),bowl_game.copy(),t1,t1,t2,factor,reduced_length)
            bat_game_2,bowl_game_2 = toss_effects(bat_game.copy(),bowl_game.copy(),t2,t1,t2,factor,reduced_length)
            bat_game['usage'] = (bat_game_1['usage']+bat_game_2['usage'])/2
            bowl_game['usage'] = (bowl_game_1['usage']+bowl_game_2['usage'])/2
        
    bowl_game['SR'] = 1/bowl_game['wickets/ball']
    
    field_game = pd.DataFrame(columns=['fielder','season','team','keeper','catches','stumpings','run outs'])
    unique_names = bat_game['batsman'].to_list() + bowl_game['bowler'].to_list()
    unique_names = list(set(unique_names))
    #field_game['fielder'] = bat_game['batsman']
    #field_game['season'] = bat_game['season']
    #field_game['team'] = bat_game['team']
    field_game['fielder'] = unique_names
    for x in unique_names:
        try:
            field_game.loc[field_game['fielder']==x,'team'] = bat_game.loc[bat_game['batsman']==x,'team'].values[0]
            field_game.loc[field_game['fielder']==x,'season'] = bat_game.loc[bat_game['batsman']==x,'season'].values[0]
        except IndexError:
            field_game.loc[field_game['fielder']==x,'team'] = bowl_game.loc[bowl_game['bowler']==x,'team'].values[0]
            field_game.loc[field_game['fielder']==x,'season'] = bowl_game.loc[bowl_game['bowler']==x,'season'].values[0]
    field_game['keeper'] = 0
    field_game['catches'] = 0
    field_game['stumpings'] = 0
    field_game['run outs'] = 0
    field_game['xPts'] = 0
    field_game.loc[field_game['fielder'].isin(designated_keeper_list), 'keeper'] = 1
    
    t1_wickets = (bat_game.loc[bat_game['team']==t2,'usage'] * bat_game.loc[bat_game['team']==t2,'wickets/ball']*factor*120).sum()
    t2_wickets = (bat_game.loc[bat_game['team']==t1,'usage'] * bat_game.loc[bat_game['team']==t1,'wickets/ball']*factor*120).sum()
    if(factor == 11.25):
        field_game.loc[(field_game['keeper']==0)&(field_game['team']==t1),'catches'] = t1_wickets*dismissals.loc[dismissals['wicket type'] == 'caught fielder','tests'].values[0]/(sum(field_game['team'] == t1)-1)
        field_game.loc[(field_game['keeper']==0)&(field_game['team']==t2),'catches'] = t2_wickets*dismissals.loc[dismissals['wicket type'] == 'caught fielder','tests'].values[0]/(sum(field_game['team'] == t2)-1)
        field_game.loc[(field_game['keeper']==1)&(field_game['team']==t1),'catches'] = t1_wickets*dismissals.loc[dismissals['wicket type'] == 'caught wk','tests'].values[0]
        field_game.loc[(field_game['keeper']==1)&(field_game['team']==t2),'catches'] = t2_wickets*dismissals.loc[dismissals['wicket type'] == 'caught wk','tests'].values[0]
        field_game.loc[field_game['team']==t1,'stumpings'] = t1_wickets*dismissals.loc[dismissals['wicket type'] == 'stumped','tests'].values[0]*field_game['keeper']
        field_game.loc[field_game['team']==t2,'stumpings'] = t2_wickets*dismissals.loc[dismissals['wicket type'] == 'stumped','tests'].values[0]*field_game['keeper']
        field_game.loc[field_game['team']==t1,'run outs'] = t1_wickets*dismissals.loc[dismissals['wicket type'] == 'run out','tests'].values[0]/(sum(field_game['team'] == t1))
        field_game.loc[field_game['team']==t2,'run outs'] = t2_wickets*dismissals.loc[dismissals['wicket type'] == 'run out','tests'].values[0]/(sum(field_game['team'] == t2))
    elif(factor == 2.5):
        field_game.loc[(field_game['keeper']==0)&(field_game['team']==t1),'catches'] = t1_wickets*dismissals.loc[dismissals['wicket type'] == 'caught fielder','odi'].values[0]/(sum(field_game['team'] == t1)-1)
        field_game.loc[(field_game['keeper']==0)&(field_game['team']==t2),'catches'] = t2_wickets*dismissals.loc[dismissals['wicket type'] == 'caught fielder','odi'].values[0]/(sum(field_game['team'] == t2)-1)
        field_game.loc[(field_game['keeper']==1)&(field_game['team']==t1),'catches'] = t1_wickets*dismissals.loc[dismissals['wicket type'] == 'caught wk','odi'].values[0]
        field_game.loc[(field_game['keeper']==1)&(field_game['team']==t2),'catches'] = t2_wickets*dismissals.loc[dismissals['wicket type'] == 'caught wk','odi'].values[0]
        field_game.loc[field_game['team']==t1,'stumpings'] = t1_wickets*dismissals.loc[dismissals['wicket type'] == 'stumped','odi'].values[0]*field_game['keeper']
        field_game.loc[field_game['team']==t2,'stumpings'] = t2_wickets*dismissals.loc[dismissals['wicket type'] == 'stumped','odi'].values[0]*field_game['keeper']
        field_game.loc[field_game['team']==t1,'run outs'] = t1_wickets*dismissals.loc[dismissals['wicket type'] == 'run out','odi'].values[0]/(sum(field_game['team'] == t1))
        field_game.loc[field_game['team']==t2,'run outs'] = t2_wickets*dismissals.loc[dismissals['wicket type'] == 'run out','odi'].values[0]/(sum(field_game['team'] == t2))
    else:
        field_game.loc[(field_game['keeper']==0)&(field_game['team']==t1),'catches'] = t1_wickets*dismissals.loc[dismissals['wicket type'] == 'caught fielder','t20'].values[0]/(sum(field_game['team'] == t1)-1)
        field_game.loc[(field_game['keeper']==0)&(field_game['team']==t2),'catches'] = t2_wickets*dismissals.loc[dismissals['wicket type'] == 'caught fielder','t20'].values[0]/(sum(field_game['team'] == t2)-1)
        field_game.loc[(field_game['keeper']==1)&(field_game['team']==t1),'catches'] = t1_wickets*dismissals.loc[dismissals['wicket type'] == 'caught wk','t20'].values[0]
        field_game.loc[(field_game['keeper']==1)&(field_game['team']==t2),'catches'] = t2_wickets*dismissals.loc[dismissals['wicket type'] == 'caught wk','t20'].values[0]
        field_game.loc[field_game['team']==t1,'stumpings'] = t1_wickets*dismissals.loc[dismissals['wicket type'] == 'stumped','t20'].values[0]*field_game['keeper']
        field_game.loc[field_game['team']==t2,'stumpings'] = t2_wickets*dismissals.loc[dismissals['wicket type'] == 'stumped','t20'].values[0]*field_game['keeper']
        field_game.loc[field_game['team']==t1,'run outs'] = t1_wickets*dismissals.loc[dismissals['wicket type'] == 'run out','t20'].values[0]/(sum(field_game['team'] == t1))
        field_game.loc[field_game['team']==t2,'run outs'] = t2_wickets*dismissals.loc[dismissals['wicket type'] == 'run out','t20'].values[0]/(sum(field_game['team'] == t2))
    
    return (bat_game,bowl_game,field_game,summary,t1,t2,s1,s2,c1,c2,league_avg,w_avg)

print("")
def gw_projection(a,b,input_file1,input_file,factor,v,tbf,reduced_length):
    
    (bat_game,bowl_game,field_game,summary,t1,t2,s1,s2,c1,c2,league_avg,w_avg) = base_calculations(a,b,input_file1,input_file,factor,v,tbf,reduced_length)
    
    dismissals = dismissal_proportions()
    
    bat_game['xPts'] = 120*reduced_length*factor*bat_game['usage']*(bat_game['runs/ball']+bat_game['4s/ball']+2*bat_game['6s/ball'])
    #bat_game['xPts'] = bat_game['xPts'] - 2*bat_game['wickets/ball']

    bat_game = bat_game.drop(['pp usage'],axis=1)
    bat_game = bat_game.drop(['mid usage'],axis=1)
    bat_game = bat_game.drop(['setup usage'],axis=1)
    bat_game = bat_game.drop(['death usage'],axis=1)
    bowl_game = bowl_game.drop(['pp usage'],axis=1)
    bowl_game = bowl_game.drop(['mid usage'],axis=1)
    bowl_game = bowl_game.drop(['setup usage'],axis=1)
    bowl_game = bowl_game.drop(['death usage'],axis=1)    
    
    SR = get_truncated_normal(mean=bat_game['SR'], sd=bat_game['SR']*0.36, low=0, upp=600)
    bf = get_truncated_normal(mean=bat_game['usage']*120*reduced_length*factor, sd=bat_game['usage']*120*reduced_length*0.6, low=0, upp=120*reduced_length*factor)
    rs = get_truncated_normal(mean=bat_game['runs/ball']*bat_game['usage']*120*reduced_length*factor, sd=bat_game['runs/ball']*bat_game['usage']*120*reduced_length*factor*0.6, low=0, upp=250*factor)
    ECON = get_truncated_normal(mean=bowl_game['ECON'], sd=bowl_game['ECON']*0.36, low=0, upp=36)
    wkts = get_truncated_normal(mean=bowl_game['wickets/ball']*bowl_game['usage']*120*reduced_length*factor, sd=bowl_game['wickets/ball']*bowl_game['usage']*120*reduced_length*factor*1.6, low=0, upp=10)
    catches = get_truncated_normal(mean=field_game['catches'], sd=field_game['catches'], low=0, upp=10)
    
    field_game['xPts'] = 8*field_game['catches'] + 12*field_game['stumpings'] + 12*field_game['run outs']
    
    if(factor != 11.25):
        bowl_game['usage'] = np.minimum(bowl_game['usage'],0.2)
        field_game['xPts'] += 4*(1-catches.cdf(3))
        
        if(factor == 1):
            b_lbw = (dismissals.loc[dismissals['wicket type']=='bowled','t20'].values[0] + dismissals.loc[dismissals['wicket type']=='lbw','t20'].values[0])/(sum(dismissals['t20'])-dismissals.loc[dismissals['wicket type']=='run out','t20'].values[0])
            bowl_game['xPts'] = 120*reduced_length*factor*bowl_game['usage']*bowl_game['wickets/ball']*(25*(1-b_lbw)+33*b_lbw)
            
            bat_game['xPts'] = bat_game['xPts'] - 2*rs.cdf(1)*(1-bf.cdf(30))
            bat_game['xPts'] = bat_game['xPts'] + 6*(1-SR.cdf(170))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] + 4*(SR.cdf(170)-SR.cdf(150))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] + 2*(SR.cdf(150)-SR.cdf(130))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] - 2*(SR.cdf(70)-SR.cdf(60))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] - 4*(SR.cdf(60)-SR.cdf(50))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] - 6*(SR.cdf(50))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] + 4*(rs.cdf(50)-rs.cdf(30))
            bat_game['xPts'] = bat_game['xPts'] + 8*(rs.cdf(100)-rs.cdf(50))
            bat_game['xPts'] = bat_game['xPts'] + 16*(1-rs.cdf(100))
            bowl_game['xPts'] = bowl_game['xPts'] + 12*np.power(bowl_game['dots/ball'],6)*20*bowl_game['usage']*factor
            bowl_game['xPts'] = bowl_game['xPts'] + 6*(ECON.cdf(5))
            bowl_game['xPts'] = bowl_game['xPts'] + 4*(ECON.cdf(6)-ECON.cdf(5))
            bowl_game['xPts'] = bowl_game['xPts'] + 2*(ECON.cdf(7)-ECON.cdf(6))
            bowl_game['xPts'] = bowl_game['xPts'] - 2*(ECON.cdf(11)-ECON.cdf(10))
            bowl_game['xPts'] = bowl_game['xPts'] - 4*(ECON.cdf(12)-ECON.cdf(11))
            bowl_game['xPts'] = bowl_game['xPts'] - 6*(1-ECON.cdf(12))
            bowl_game['xPts'] = bowl_game['xPts'] + 4*(wkts.cdf(4)-wkts.cdf(3))
            bowl_game['xPts'] = bowl_game['xPts'] + 8*(wkts.cdf(5)-wkts.cdf(4))
            bowl_game['xPts'] = bowl_game['xPts'] + 16*(1-wkts.cdf(5))
            
        elif(factor == 2.5):
            b_lbw = (dismissals.loc[dismissals['wicket type']=='bowled','odi'].values[0] + dismissals.loc[dismissals['wicket type']=='lbw','odi'].values[0])/(sum(dismissals['odi'])-dismissals.loc[dismissals['wicket type']=='run out','odi'].values[0])
            bowl_game['xPts'] = 120*reduced_length*factor*bowl_game['usage']*bowl_game['wickets/ball']*(25*(1-b_lbw)+33*b_lbw)
            
            bat_game['xPts'] = bat_game['xPts'] - 3*rs.cdf(1)*(1-bf.cdf(30))
            bat_game['xPts'] = bat_game['xPts'] + 6*(1-SR.cdf(140))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] + 4*(SR.cdf(140)-SR.cdf(120))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] + 2*(SR.cdf(120)-SR.cdf(100))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] - 2*(SR.cdf(50)-SR.cdf(40))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] - 4*(SR.cdf(40)-SR.cdf(30))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] - 6*(SR.cdf(30))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] + 4*(rs.cdf(100)-rs.cdf(50))
            bat_game['xPts'] = bat_game['xPts'] + 8*(rs.cdf(200)-rs.cdf(100))
            bat_game['xPts'] = bat_game['xPts'] + 16*(1-rs.cdf(200))
            bowl_game['xPts'] = bowl_game['xPts'] + 12*np.power(bowl_game['dots/ball'],6)*20*bowl_game['usage']*factor
            bowl_game['xPts'] = bowl_game['xPts'] + 6*(ECON.cdf(2.5))
            bowl_game['xPts'] = bowl_game['xPts'] + 4*(ECON.cdf(3.5)-ECON.cdf(2.5))
            bowl_game['xPts'] = bowl_game['xPts'] + 2*(ECON.cdf(4.5)-ECON.cdf(3.5))
            bowl_game['xPts'] = bowl_game['xPts'] - 2*(ECON.cdf(8)-ECON.cdf(7))
            bowl_game['xPts'] = bowl_game['xPts'] - 4*(ECON.cdf(9)-ECON.cdf(8))
            bowl_game['xPts'] = bowl_game['xPts'] - 6*(1-ECON.cdf(9))
            bowl_game['xPts'] = bowl_game['xPts'] + 0*(wkts.cdf(4)-wkts.cdf(3))
            bowl_game['xPts'] = bowl_game['xPts'] + 4*(wkts.cdf(5)-wkts.cdf(4))
            bowl_game['xPts'] = bowl_game['xPts'] + 8*(1-wkts.cdf(5))
            
        elif(factor == 5/6):
            b_lbw = (dismissals.loc[dismissals['wicket type']=='bowled','t20'].values[0] + dismissals.loc[dismissals['wicket type']=='lbw','t20'].values[0])/(sum(dismissals['t20'])-dismissals.loc[dismissals['wicket type']=='run out','t20'].values[0])
            bowl_game['xPts'] = 120*reduced_length*factor*bowl_game['usage']*bowl_game['wickets/ball']*(25*(1-b_lbw)+33*b_lbw)
            
            bat_game['xPts'] = bat_game['xPts'] + 5*(rs.cdf(50)-rs.cdf(30))
            bat_game['xPts'] = bat_game['xPts'] + 10*(rs.cdf(100)-rs.cdf(50))
            bat_game['xPts'] = bat_game['xPts'] + 20*(1-rs.cdf(100))
            bowl_game['xPts'] = bowl_game['xPts'] + 16*np.power(bowl_game['dots/ball'],6)*20*bowl_game['usage']*factor
            bowl_game['xPts'] = bowl_game['xPts'] + 5*(wkts.cdf(4)-wkts.cdf(3))
            bowl_game['xPts'] = bowl_game['xPts'] + 10*(wkts.cdf(5)-wkts.cdf(4))
            bowl_game['xPts'] = bowl_game['xPts'] + 20*(1-wkts.cdf(5))
            
    else:
        b_lbw = (dismissals.loc[dismissals['wicket type']=='bowled','tests'].values[0] + dismissals.loc[dismissals['wicket type']=='lbw','tests'].values[0])/(sum(dismissals['tests'])-dismissals.loc[dismissals['wicket type']=='run out','tests'].values[0])
        bowl_game['xPts'] = 120*reduced_length*factor*bowl_game['usage']*bowl_game['wickets/ball']*(25*(1-b_lbw)+33*b_lbw)
        
        bat_game['xPts'] = bat_game['xPts'] - 4*rs.cdf(1)*(1-bf.cdf(30))
        bowl_game['xPts'] = 120*reduced_length*factor*bowl_game['usage']*bowl_game['wickets/ball']*19        
        bat_game['xPts'] = bat_game['xPts'] + 4*(rs.cdf(100)-rs.cdf(50))
        bat_game['xPts'] = bat_game['xPts'] + 8*(rs.cdf(200)-rs.cdf(100))
        bat_game['xPts'] = bat_game['xPts'] + 16*(rs.cdf(300)-rs.cdf(200))
        bat_game['xPts'] = bat_game['xPts'] + 24*(1-rs.cdf(300))
        bowl_game['xPts'] = bowl_game['xPts'] + 4*(wkts.cdf(5)-wkts.cdf(4))
        bowl_game['xPts'] = bowl_game['xPts'] + 8*(1-wkts.cdf(5))
    
    runs_t1,runs_t2 = game_results(bat_game,factor,t1,t2,s1,s2,c1,c2,league_avg,summary,reduced_length)
    return (bat_game,bowl_game,field_game,summary,runs_t1,runs_t2)

def cricdraft_projection(a,b,input_file1,input_file,factor,v,tbf,reduced_length):
    
    (bat_game,bowl_game,field_game,summary,t1,t2,s1,s2,c1,c2,league_avg,w_avg) = base_calculations(a,b,input_file1,input_file,factor,v,tbf,reduced_length)
    
    bat_game = bat_game.drop(['pp usage'],axis=1)
    bat_game = bat_game.drop(['mid usage'],axis=1)
    bat_game = bat_game.drop(['setup usage'],axis=1)
    bat_game = bat_game.drop(['death usage'],axis=1)
    bowl_game = bowl_game.drop(['pp usage'],axis=1)
    bowl_game = bowl_game.drop(['mid usage'],axis=1)
    bowl_game = bowl_game.drop(['setup usage'],axis=1)
    bowl_game = bowl_game.drop(['death usage'],axis=1)
    
    SR = get_truncated_normal(mean=bat_game['SR'], sd=bat_game['SR']*0.36, low=0, upp=600)
    bf = get_truncated_normal(mean=bat_game['usage']*120*reduced_length*factor, sd=bat_game['usage']*120*reduced_length*0.6, low=0, upp=120*reduced_length*factor)
    rs = get_truncated_normal(mean=bat_game['runs/ball']*bat_game['usage']*120*reduced_length*factor, sd=bat_game['runs/ball']*bat_game['usage']*120*reduced_length*factor*0.6, low=0, upp=250*factor)
    ECON = get_truncated_normal(mean=bowl_game['ECON'], sd=bowl_game['ECON']*0.36, low=0, upp=42)
    wkts = get_truncated_normal(mean=bowl_game['wickets/ball']*bowl_game['usage']*120*reduced_length*factor, sd=bowl_game['wickets/ball']*bowl_game['usage']*120*reduced_length*factor*1.6, low=0, upp=10)
    bb = get_truncated_normal(mean=bowl_game['usage']*120*reduced_length*factor, sd=bowl_game['usage']*120*reduced_length*0.2, low=0, upp=0.2*120*reduced_length*factor)
    catches = get_truncated_normal(mean=field_game['catches'], sd=field_game['catches'], low=0, upp=10)
    
    field_game['xPts'] = 10*field_game['catches'] + 25*field_game['stumpings'] + 25*field_game['run outs']
    
    if(factor != 11.25):
        bat_game['xPts'] = 120*reduced_length*factor*bat_game['usage']*(bat_game['runs/ball']+3*bat_game['6s/ball'])
        bowl_game['usage'] = np.minimum(bowl_game['usage'],0.2)
        bowl_game['xPts'] = 120*reduced_length*factor*bowl_game['usage']*bowl_game['wickets/ball']*30
        #bat_game['xPts'] = bat_game['xPts'] - 25*rs.cdf(1)*(1-bf.cdf(4))
        bat_bonus,bowl_bonus = cricdraft_bonus(SR,bf,rs,ECON,wkts,bb)
        if(factor <= 1):
            #multiplication of 2 gaussians is not a gaussian !
            bat_game['xPts'] += bat_bonus
            bowl_game['xPts'] += 25*np.power(bowl_game['dots/ball'],6)*20*bowl_game['usage']*factor
            bowl_game['xPts'] += bowl_bonus
            
        elif(factor == 2.5):
            bat_game['xPts'] = bat_game['xPts'] + 2.5*bat_game['runs/ball']*bat_game['usage']*120*reduced_length*factor*(1-SR.cdf(200))*(1-bf.cdf(24))
            bat_game['xPts'] = bat_game['xPts'] + 2*bat_game['runs/ball']*bat_game['usage']*120*reduced_length*factor*(SR.cdf(200)-SR.cdf(170))*(1-bf.cdf(24))
            bat_game['xPts'] = bat_game['xPts'] + 1.5*bat_game['runs/ball']*bat_game['usage']*120*reduced_length*factor*(SR.cdf(170)-SR.cdf(145))*(1-bf.cdf(24))
            bat_game['xPts'] = bat_game['xPts'] + bat_game['runs/ball']*bat_game['usage']*120*reduced_length*factor*(SR.cdf(145)-SR.cdf(120))*(1-bf.cdf(24))
            bat_game['xPts'] = bat_game['xPts'] + 0.5*bat_game['runs/ball']*bat_game['usage']*120*reduced_length*factor*(SR.cdf(120)-SR.cdf(95))*(1-bf.cdf(24))
            bat_game['xPts'] = bat_game['xPts'] - 5*(SR.cdf(80)-SR.cdf(65))*(1-bf.cdf(24))
            bat_game['xPts'] = bat_game['xPts'] - 10*(SR.cdf(65)-SR.cdf(50))*(1-bf.cdf(24))
            bat_game['xPts'] = bat_game['xPts'] - 15*(SR.cdf(50)-SR.cdf(40))*(1-bf.cdf(24))
            bat_game['xPts'] = bat_game['xPts'] - 20*(SR.cdf(40))*(1-bf.cdf(24))
            bat_game['xPts'] = bat_game['xPts'] + 25*(rs.cdf(100)-rs.cdf(50))
            bat_game['xPts'] = bat_game['xPts'] + 50*(rs.cdf(150)-rs.cdf(100))
            #bat_game['xPts'] = bat_game['xPts'] + 100*(1-rs.cdf(150))
            bowl_game['xPts'] = bowl_game['xPts'] + 15*np.power(bowl_game['dots/ball'],6)*20*bowl_game['usage']*factor
            bowl_game['xPts'] = bowl_game['xPts'] + 100*(ECON.cdf(2.5))*(1-bb.cdf(24))
            bowl_game['xPts'] = bowl_game['xPts'] + 70*(ECON.cdf(3.5)-ECON.cdf(2.5))*(1-bb.cdf(24))
            bowl_game['xPts'] = bowl_game['xPts'] + 45*(ECON.cdf(4.25)-ECON.cdf(3.5))*(1-bb.cdf(24))
            bowl_game['xPts'] = bowl_game['xPts'] + 25*(ECON.cdf(4.75)-ECON.cdf(4.25))*(1-bb.cdf(24))
            bowl_game['xPts'] = bowl_game['xPts'] + 10*(ECON.cdf(5.25)-ECON.cdf(4.75))*(1-bb.cdf(24))
            bowl_game['xPts'] = bowl_game['xPts'] - 5*(ECON.cdf(6.5)-ECON.cdf(5.75))*(1-bb.cdf(24))
            bowl_game['xPts'] = bowl_game['xPts'] - 10*(ECON.cdf(8)-ECON.cdf(6.5))*(1-bb.cdf(24))
            bowl_game['xPts'] = bowl_game['xPts'] - 15*(ECON.cdf(10.5)-ECON.cdf(8))*(1-bb.cdf(24))
            bowl_game['xPts'] = bowl_game['xPts'] - 20*(1-ECON.cdf(10.5))*(1-bb.cdf(24))
            bowl_game['xPts'] = bowl_game['xPts'] + 15*(wkts.cdf(4)-wkts.cdf(3))
            bowl_game['xPts'] = bowl_game['xPts'] + 25*(wkts.cdf(5)-wkts.cdf(4))
            bowl_game['xPts'] = bowl_game['xPts'] + 50*(1-wkts.cdf(5))
            #bowl_game['xPts'] = bowl_game['xPts'] + 75*(wkts.cdf(7)-wkts.cdf(6))
            #bowl_game['xPts'] = bowl_game['xPts'] + 100*(1-wkts.cdf(7))
            
    else:
        bowl_game['xPts'] = 120*reduced_length*factor*bowl_game['usage']*bowl_game['wickets/ball']*19        
        bat_game['xPts'] = bat_game['xPts'] + 4*(rs.cdf(100)-rs.cdf(50))
        bat_game['xPts'] = bat_game['xPts'] + 8*(rs.cdf(200)-rs.cdf(100))
        bat_game['xPts'] = bat_game['xPts'] + 16*(rs.cdf(300)-rs.cdf(200))
        bat_game['xPts'] = bat_game['xPts'] + 24*(1-rs.cdf(300))
        bowl_game['xPts'] = bowl_game['xPts'] + 4*(wkts.cdf(5)-wkts.cdf(4))
        bowl_game['xPts'] = bowl_game['xPts'] + 8*(1-wkts.cdf(5))
    
    runs_t1,runs_t2 = game_results(bat_game,factor,t1,t2,s1,s2,c1,c2,league_avg,summary,reduced_length)
    return (bat_game,bowl_game,field_game,summary,runs_t1,runs_t2)

def ex22_projection(a,b,input_file1,input_file,factor,v,tbf,reduced_length):
    
    (bat_game,bowl_game,field_game,summary,t1,t2,s1,s2,c1,c2,league_avg,w_avg) = base_calculations(a,b,input_file1,input_file,factor,v,tbf,reduced_length)

    bat_game = bat_game.drop(['pp usage'],axis=1)
    bat_game = bat_game.drop(['mid usage'],axis=1)
    bat_game = bat_game.drop(['setup usage'],axis=1)
    bat_game = bat_game.drop(['death usage'],axis=1)
    bowl_game = bowl_game.drop(['pp usage'],axis=1)
    bowl_game = bowl_game.drop(['mid usage'],axis=1)
    bowl_game = bowl_game.drop(['setup usage'],axis=1)
    bowl_game = bowl_game.drop(['death usage'],axis=1)
    
    bat_game['xPts'] = 0.0; bowl_game['xPts'] = 0.0
    rs = get_truncated_normal(mean=bat_game['runs/ball']*bat_game['usage']*120*reduced_length*factor, sd=bat_game['runs/ball']*bat_game['usage']*120*reduced_length*factor*0.6, low=0, upp=250*factor)
    wkts = get_truncated_normal(mean=bowl_game['wickets/ball']*bowl_game['usage']*120*reduced_length*factor, sd=bowl_game['wickets/ball']*bowl_game['usage']*120*reduced_length*factor*1.6, low=0, upp=10)
    
    field_game['xPts'] = 8*field_game['catches'] + 12*field_game['stumpings'] + 12*field_game['run outs']
    
    if(factor == 1):        
        bat_game['xPts'] = bat_game['xPts']+(bat_game['usage']*(bat_game['runs/ball']+bat_game['4s/ball']+2*bat_game['6s/ball'])*120*reduced_length*factor)
        bat_game['xPts'] += 7*(rs.cdf(100)-rs.cdf(50))
        bat_game['xPts'] += 15*(1-rs.cdf(100))
        bowl_game['usage'] = np.minimum(bowl_game['usage'],0.2)
        bowl_game['xPts'] += 20*bowl_game['usage']*bowl_game['wickets/ball']*120*reduced_length*factor
        bowl_game['xPts'] += 5*(wkts.cdf(4)-wkts.cdf(3))
        bowl_game['xPts'] += 8*(1-wkts.cdf(4))
        bowl_game['xPts'] += 7*np.power(bowl_game['dots/ball'],6)*20*bowl_game['usage']*factor
        
    elif(factor == 2.5):
        bat_game['xPts'] += bat_game['usage']*(bat_game['runs/ball']+bat_game['4s/ball']+2*bat_game['6s/ball'])*120*reduced_length*factor
        bat_game['xPts'] += 5*(rs.cdf(100)-rs.cdf(50))
        bat_game['xPts'] += 8*(1-rs.cdf(100))
        bowl_game['usage'] = np.minimum(bowl_game['usage'],0.2)
        bowl_game['xPts'] += 25*bowl_game['usage']*bowl_game['wickets/ball']*120*reduced_length*factor
        bowl_game['xPts'] += 5*(wkts.cdf(5)-wkts.cdf(4))
        bowl_game['xPts'] += 8*(1-wkts.cdf(5))
        bowl_game['xPts'] += 5*np.power(bowl_game['dots/ball'],6)*20*bowl_game['usage']*factor
        
    elif(factor == 5/6):         
        bat_game['xPts'] += bat_game['usage']*(bat_game['runs/ball']+bat_game['4s/ball']+2*bat_game['6s/ball'])*120*reduced_length*factor
        bat_game['xPts'] += 8*(1-rs.cdf(50))
        bowl_game['usage'] = np.minimum(bowl_game['usage'],0.2)
        bowl_game['xPts'] += 20*bowl_game['usage']*bowl_game['wickets/ball']*120*reduced_length*factor
        bowl_game['xPts'] += 5*(wkts.cdf(4)-wkts.cdf(3))
        bowl_game['xPts'] += 8*(1-wkts.cdf(4))
        
    else: #test match
        wkts = get_truncated_normal(mean=bowl_game['wickets/ball']*(bowl_game['usage']/2)*120*reduced_length*factor, sd=bowl_game['wickets/ball']*bowl_game['usage']*120*reduced_length*factor*1.6, low=0, upp=10)
        bat_game['xPts'] += bat_game['usage']*(0.5*bat_game['runs/ball']+bat_game['4s/ball']+3*bat_game['6s/ball'])*120*reduced_length*factor
        bat_game['xPts'] += 5*(rs.cdf(100)-rs.cdf(50))
        bat_game['xPts'] += 8*(rs.cdf(200)-rs.cdf(100))
        bat_game['xPts'] += 15*(1-rs.cdf(200))
        bowl_game['xPts'] += 12*bowl_game['usage']*bowl_game['wickets/ball']*120*reduced_length*factor
        bowl_game['xPts'] += 10*(1-wkts.cdf(5))
        bowl_game['xPts'] += 1*np.power(bowl_game['dots/ball'],6)*20*bowl_game['usage']*factor
     
    runs_t1,runs_t2 = game_results(bat_game,factor,t1,t2,s1,s2,c1,c2,league_avg,summary,reduced_length)
    return (bat_game,bowl_game,field_game,summary,runs_t1,runs_t2)

def coversoff_projection(a,b,input_file1,input_file,factor,v,tbf,reduced_length):
    
    (bat_game,bowl_game,field_game,summary,t1,t2,s1,s2,c1,c2,league_avg,w_avg) = base_calculations(a,b,input_file1,input_file,factor,v,tbf,reduced_length)
        
    bat_game = bat_game.drop(['pp usage'],axis=1)
    bat_game = bat_game.drop(['mid usage'],axis=1)
    bat_game = bat_game.drop(['setup usage'],axis=1)
    bat_game = bat_game.drop(['death usage'],axis=1)
    #initialize points as 0
    bat_game['xPts'] = 0.0; bowl_game['xPts'] = 0.0
    SR = get_truncated_normal(mean=bat_game['SR'], sd=bat_game['SR']*0.36, low=0, upp=600)
    bf = get_truncated_normal(mean=bat_game['usage']*120*reduced_length*factor, sd=bat_game['usage']*120*reduced_length*0.6, low=0, upp=120*reduced_length*factor)
    rs = get_truncated_normal(mean=bat_game['runs/ball']*bat_game['usage']*120*reduced_length*factor, sd=bat_game['runs/ball']*bat_game['usage']*120*reduced_length*factor*0.6, low=0, upp=250*factor)
    ECON = get_truncated_normal(mean=bowl_game['ECON'], sd=bowl_game['ECON']*0.36, low=0, upp=36)
    wkts = get_truncated_normal(mean=bowl_game['wickets/ball']*bowl_game['usage']*120*reduced_length*factor, sd=bowl_game['wickets/ball']*bowl_game['usage']*120*reduced_length*factor*1.6, low=0, upp=10)
    bb = get_truncated_normal(mean=bowl_game['usage']*120*reduced_length*factor, sd=bowl_game['usage']*120*reduced_length*0.2, low=0, upp=0.2*120*reduced_length*factor)   
    
    if(factor < 1):
        bat_game['xPts'] = bat_game['usage']*bat_game['SR']
    elif(factor == 1):
        (bat_bonus,bowl_bonus) = coversoff_bonus(SR,bf,rs,ECON,wkts,bb,factor)
        #runs 4s and 6s have pts
        bat_game['xPts'] = 120*reduced_length*factor*bat_game['usage']*(bat_game['runs/ball']+10*bat_game['6s/ball']+4*bat_game['4s/ball'])
        #SR bonuses
        bat_game['xPts'] += bat_bonus
    elif(factor == 2.5):
        #runs 4s and 6s have pts
        bat_game['xPts'] = 120*reduced_length*factor*bat_game['usage']*(bat_game['runs/ball']+3*bat_game['6s/ball']+2*bat_game['4s/ball'])
        #duck        
        bat_game['xPts'] = bat_game['xPts'] - 20*rs.cdf(1)*(1-bf.cdf(30))
        #less than x runs -25
        bat_game['xPts'] = bat_game['xPts'] - 5*(rs.cdf(5)-rs.cdf(1))*(1-bf.cdf(30))
        #run rate bonus
        #bat_game['xPts'] = bat_game['xPts'] + 120*reduced_length*factor*bat_game['usage']*bat_game['runs/ball'] - (4/3)*(bat_game['usage']*factor*120*reduced_length)
        #SR bonuses
        bat_game['xPts'] = bat_game['xPts'] + 30*(1-SR.cdf(150))*(1-bf.cdf(20))
        bat_game['xPts'] = bat_game['xPts'] + 20*(SR.cdf(150)-SR.cdf(125))*(1-bf.cdf(20))
        bat_game['xPts'] = bat_game['xPts'] - 10*(SR.cdf(125)-SR.cdf(100))*(1-bf.cdf(20))
        bat_game['xPts'] = bat_game['xPts'] + 10*(SR.cdf(100)-SR.cdf(75))*(1-bf.cdf(20))
        bat_game['xPts'] = bat_game['xPts'] - 20*(SR.cdf(75)-SR.cdf(50))*(1-bf.cdf(20))
        bat_game['xPts'] = bat_game['xPts'] - 30*(SR.cdf(50)-SR.cdf(25))*(1-bf.cdf(20))
        bat_game['xPts'] = bat_game['xPts'] - 50*(SR.cdf(25))*(1-bf.cdf(20))
        #milestone bonuses
        bat_game['xPts'] = bat_game['xPts'] + 10*(rs.cdf(50)-rs.cdf(30))
        bat_game['xPts'] = bat_game['xPts'] + 20*(rs.cdf(75)-rs.cdf(50))
        bat_game['xPts'] = bat_game['xPts'] + 30*(rs.cdf(100)-rs.cdf(75))
        bat_game['xPts'] = bat_game['xPts'] + 50*(rs.cdf(150)-rs.cdf(100))
        bat_game['xPts'] = bat_game['xPts'] + 75*(1-rs.cdf(150))
        
        
    #bowler cant bowl more than 20% of the overs except in tests
    if(factor!=11.25):bowl_game['usage'] = np.minimum(bowl_game['usage'],0.2)
    bowl_game = bowl_game.drop(['pp usage'],axis=1)
    bowl_game = bowl_game.drop(['mid usage'],axis=1)
    bowl_game = bowl_game.drop(['setup usage'],axis=1)
    bowl_game = bowl_game.drop(['death usage'],axis=1)
    
    if(factor < 1):
        bowl_game['xPts'] = bowl_game['usage']*bowl_game['wickets/ball']*100*20
    if(factor == 1):
        #wicket is 35 pts
        bowl_game['xPts'] = 120*reduced_length*factor*bowl_game['usage']*bowl_game['wickets/ball']*35
        #dot balls pts
        bowl_game['xPts'] += 120*reduced_length*factor*bowl_game['usage']*bowl_game['dots/ball']*5
        #ECON bonuses
        bowl_game['xPts'] += bowl_bonus 
        #maiden over
        bowl_game['xPts'] = bowl_game['xPts'] + 100*np.power(bowl_game['dots/ball'],6)*20*bowl_game['usage']*factor
        
    elif(factor == 2.5):
        #wicket is 30 pts
        bowl_game['xPts'] = 120*reduced_length*factor*bowl_game['usage']*bowl_game['wickets/ball']*35
        #dot balls pts
        #bowl_game['xPts'] += 120*reduced_length*factor*bowl_game['usage']*bowl_game['dots/ball']*5
        #ECON bonuses
        bowl_game['xPts'] = bowl_game['xPts'] - 40*(1-ECON.cdf(9))*(1-bb.cdf(18))
        bowl_game['xPts'] = bowl_game['xPts'] - 20*(ECON.cdf(9)-ECON.cdf(9))*(1-bb.cdf(18))
        bowl_game['xPts'] = bowl_game['xPts'] - 10*(ECON.cdf(7)-ECON.cdf(6))*(1-bb.cdf(18))
        bowl_game['xPts'] = bowl_game['xPts'] + 10*(ECON.cdf(6)-ECON.cdf(5))*(1-bb.cdf(18))
        bowl_game['xPts'] = bowl_game['xPts'] + 20*(ECON.cdf(5)-ECON.cdf(4))*(1-bb.cdf(18))
        bowl_game['xPts'] = bowl_game['xPts'] + 30*(ECON.cdf(4)-ECON.cdf(3))*(1-bb.cdf(18))
        bowl_game['xPts'] = bowl_game['xPts'] + 50*(ECON.cdf(3))*(1-bb.cdf(18))
        #milestone bonuses
        bowl_game['xPts'] = bowl_game['xPts'] + 10*(wkts.cdf(3)-wkts.cdf(2))
        bowl_game['xPts'] = bowl_game['xPts'] + 30*(wkts.cdf(4)-wkts.cdf(3))
        bowl_game['xPts'] = bowl_game['xPts'] + 50*(wkts.cdf(5)-wkts.cdf(4))
        bowl_game['xPts'] = bowl_game['xPts'] + 70*(wkts.cdf(6)-wkts.cdf(5))
        bowl_game['xPts'] = bowl_game['xPts'] + 100*(1-wkts.cdf(6))
        #maiden over
        bowl_game['xPts'] = bowl_game['xPts'] + 15*np.power(bowl_game['dots/ball'],6)*20*bowl_game['usage']*factor
    

    runs_t1,runs_t2 = game_results(bat_game,factor,t1,t2,s1,s2,c1,c2,league_avg,summary,reduced_length)
    return (bat_game,bowl_game,field_game,summary,runs_t1,runs_t2)

def game_results(bat_game,factor,t1,t2,s1,s2,c1,c2,league_avg,summary,reduced_length):
    wickets_t1 = round((bat_game.loc[bat_game['team']==t1,'usage']*bat_game.loc[bat_game['team']==t1,'wickets/ball']).sum()*120*factor,1)
    wickets_t2 = round((bat_game.loc[bat_game['team']==t2,'usage']*bat_game.loc[bat_game['team']==t2,'wickets/ball']).sum()*120*factor,1)
    if(factor == 5/6): factor = 1
    runs_t1 = round(s1*c2*league_avg*(bat_game.loc[bat_game['team']==t1,'usage'].sum()),2)
    runs_t2 = round(s2*c1*league_avg*(bat_game.loc[bat_game['team']==t2,'usage'].sum()),2)
    
    gamma = 1
    if(factor == 11.25): gamma = 1
    elif(factor == 2.5): gamma = 4.8
    else: gamma = 6.4
    win_t1 = pow(t_runs_t1,gamma)/(pow(t_runs_t1,gamma)+pow(t_runs_t2,gamma))
    win_t2 = pow(t_runs_t2,gamma)/(pow(t_runs_t1,gamma)+pow(t_runs_t2,gamma))
    print(t1,runs_t1,"runs for",wickets_t1,"wickets in",round(20*reduced_length*factor*bat_game.loc[bat_game['team']==t1,'usage'].sum(),2),"overs")
    print(t2,runs_t2,"runs for",wickets_t2,"wickets in",round(20*reduced_length*factor*bat_game.loc[bat_game['team']==t2,'usage'].sum(),2),"overs")
    if(factor != 11.25): print("P(win) :-",t1,round(win_t1,3),"-",round(win_t2,3),t2)
    global home_win
    home_win.append(win_t1)
    print(" ")
    return (runs_t1,runs_t2)

def cricdraft_bonus(SR,bf,rs,ECON,wkts,bb):
    c = 0; bat_total = 0; bowl_total = 0
    while(c<1000):
        try:
            bfi = bf.rvs()
            SRi = SR.rvs()
            bbi = bb.rvs()
        except ValueError:
            bfi = 0
            SRi = 0
            bbi = 0
        rsi = bfi*SRi/100
        milstonei = np.where((rsi>=50)&(rsi<100), 25, 0) + np.where((rsi>=100)&(rsi<75), 50, 0) + np.where(rsi>=150, 75, 0)
        SR_bonusi = np.where((SRi>=0)&(SRi<50), -20, 0) + np.where((SRi>=50)&(SRi<70), -15, 0) + np.where((SRi>=70)&(SRi<90), -10, 0) + \
                    np.where((SRi>=90)&(SRi<100), -5, 0) + np.where((SRi>=140)&(SRi<170), 0.5*rsi, 0) + np.where((SRi>=170)&(SRi<200), rsi, 0) + \
                    np.where((SRi>=200)&(SRi<230), 1.5*rsi, 0) + np.where(SRi>=230, 2*rsi, 0)
        qualifyi =  np.where((rsi>=16)|(bfi>=12), 1, 0)
        bat_total += milstonei + qualifyi*SR_bonusi
        #bbi = bb.rvs()
        ECONi = ECON.rvs()
        rci = bbi*ECONi/6
        wktsi = wkts.rvs()
        wickets_multiplei = np.where((wktsi>=3)&(wktsi<4), 25, 0) + np.where((wktsi>=4)&(wktsi<5), 35, 0) + np.where((wktsi>=5)&(wktsi<6), 50, 0) + \
                            np.where((wktsi>=6)&(wktsi<7), 75, 0) + np.where(wktsi>=7, 100, 0)
        ECON_bonusi = np.where((ECONi>=0)&(ECONi<4), 100, 0) + np.where((ECONi>=4)&(ECONi<5), 50, 0) + np.where((ECONi>=5)&(ECONi<6), 25, 0) + np.where((ECONi>=6)&(ECONi<7), 15, 0) + \
                      np.where((ECONi>=9)&(ECONi<10), -5, 0) +np.where((ECONi>=10)&(ECONi<12), -10, 0) + np.where((ECONi>=12)&(ECONi<14), -15, 0) + np.where(ECONi>=14, -20, 0)
        qualify2i = np.where((rci>=16)|(bbi>=12), 1, 0)
        bowl_total += wickets_multiplei + qualify2i*ECON_bonusi
        c += 1
    #print (bat_total/1000)
    #print (bowl_total/1000)
    return ((bat_total/1000),(bowl_total/1000))

def coversoff_bonus(SR,bf,rs,ECON,wkts,bb,factor):
    c = 0; bat_total = 0; bowl_total = 0
    while(c<1000):
        try:
            bfi = bf.rvs()
            SRi = SR.rvs()
            bbi = bb.rvs()
        except ValueError:
            bfi = 0
            SRi = 0
            bbi = 0
        rsi = bfi*SRi/100
        milstonei = np.where((rsi>=20)&(rsi<30), 20, 0) + np.where((rsi>=30)&(rsi<50), 40, 0) + np.where((rsi>=50)&(rsi<60), 70, 0) + \
                    np.where((rsi>=65)&(rsi<75), 120, 0) + np.where(rsi>=75, 150, 0)
        SR_bonusi = np.where((SRi>=0)&(SRi<100)&(bfi>=6), -100, 0) + np.where((SRi>=100)&(SRi<135)&(bfi>=6), -80, 0) + \
                    np.where((SRi>=140)&(SRi<160)&(rsi>=50), 60, 0) + np.where((SRi>=160)&(SRi<180)&(rsi>=40), 80, 0) + \
                    np.where((SRi>=180)&(SRi<200)&(rsi>=30), 100, 0) + np.where((SRi>=200)&(rsi>=25), 120, 0)
        qualifyi =  1
        bat_total += milstonei + qualifyi*SR_bonusi
        #bbi = bb.rvs()
        ECONi = ECON.rvs()
        rci = bbi*ECONi/(6*factor)
        wktsi = wkts.rvs()
        wickets_multiplei = np.where((wktsi>=3)&(wktsi<4), 80, 0) + np.where((wktsi>=4)&(wktsi<5), 150, 0) + np.where(wktsi>=5, 200, 0)
        ECON_bonusi = np.where((ECONi>=0)&(ECONi<4)&(bbi>=18), 150, 0) + np.where((ECONi>=4)&(ECONi<6)&(bbi>=18), 120, 0) + \
                      np.where((ECONi>=6)&(ECONi<7.5)&(bbi>=18), 100, 0) + np.where((ECONi>=7.5)&(ECONi<8)&(bbi>=24), 80, 0) + \
                      np.where((ECONi>=8.5)&(ECONi<9.5)&(bbi>=6), -80, 0) + np.where((ECONi>=9.5)&(bbi>=6), -100, 0)
        qualify2i = 1
        bowl_total += wickets_multiplei + qualify2i*ECON_bonusi
        c += 1
    #print (bat_total/1000)
    #print (bowl_total/1000)
    return ((bat_total/1000),(bowl_total/1000))

def final_projections(bat_game,bowl_game,field_game,runs_t1,runs_t2):
    names = np.unique(np.concatenate([bat_game['batsman'].values,bowl_game['bowler'].values]))
    final = [["player","team","total","bat","bowl","field","bat usage","bowl usage","+/-","batting order"]]

    for x in names:
        try : p_bat = bat_game.loc[(bat_game['batsman']==x),'team'].values[0]  
        except IndexError: p_bat = "Free Agent"
        try : p_bowl = bowl_game.loc[(bowl_game['bowler']==x),'team'].values[0]
        except IndexError: p_bowl = "Free Agent"
        if(p_bat == "Free Agent"): p_team = p_bowl
        else : p_team = p_bat
        
        bat_game = bat_game.groupby(['batsman','season','team'], as_index=False).sum()
        bowl_game = bowl_game.groupby(['bowler','season','team'], as_index=False).sum()
        field_game = field_game.groupby(['fielder','season','team'], as_index=False).sum()
        #partimers removed from bowling list, min 1 over in t20s, review this
        if(coversoff == 1 or cricdraft == 1): bowl_game.drop(bowl_game[bowl_game['usage'] < 0.025].index, inplace = True)
        else: bowl_game.drop(bowl_game[bowl_game['usage'] < 0.01].index, inplace = True)
        final.append([x,p_team,0,bat_game.loc[(bat_game['batsman']==x),'xPts'].mean(),bowl_game.loc[(bowl_game['bowler']==x),'xPts'].mean(),field_game.loc[(field_game['fielder']==x),'xPts'].mean(),bat_game.loc[(bat_game['batsman']==x),'usage'].mean(),bowl_game.loc[(bowl_game['bowler']==x),'usage'].mean(),0,bat_game.loc[(bat_game['batsman']==x),'batting_order'].mean()])
        
    final = pd.DataFrame(final)
    final.columns = final.iloc[0];final = final.drop(0)
    final = final.fillna(0)
    if(coversoff == 1): final['total'] = final['bat'] + final['bowl'] + final['field'] #+ 25
    elif(ex22 == 1 or cricdraft == 1): final['total'] = final['bat'] + final['bowl'] + final['field']
    else: final['total'] = final['bat'] + final['bowl'] + 4  + final['field']
    #final = final.sort_values(['total'], ascending=[False])
    
    for x in names:
        final.loc[final['player']==x,'bat +/-'] = 1.2*f*bat_game.loc[bat_game['batsman']==x,'usage'].sum()*(bat_game.loc[bat_game['batsman']==x,'SR'].sum()-bat_game.loc[bat_game['batsman']==x,'xSR'].sum())
        final.loc[final['player']==x,'bowl +/-'] = 20*f*bowl_game.loc[bowl_game['bowler']==x,'usage'].sum()*(bowl_game.loc[bowl_game['bowler']==x,'ECON'].sum()-bowl_game.loc[bowl_game['bowler']==x,'xECON'].sum())
    
    c=0
    while c<len(home):
        pm_t1_bat = final.loc[final['team']==home[c],'bat +/-'].sum()
        pm_t1_bowl = final.loc[final['team']==home[c],'bowl +/-'].sum()
        pm_t2_bat = final.loc[final['team']==opps[c],'bat +/-'].sum()
        pm_t2_bowl = final.loc[final['team']==opps[c],'bowl +/-'].sum()
        final.loc[final['team']==home[c],'bat +/-'] = final.loc[final['team']==home[c],'bat +/-'] + ((runs_t1[c]-game_avg[c]-pm_t1_bat) * final.loc[final['team']==home[c],'bat usage'])
        final.loc[final['team']==opps[c],'bat +/-'] = final.loc[final['team']==opps[c],'bat +/-'] + ((runs_t2[c]-game_avg[c]-pm_t2_bat) * final.loc[final['team']==opps[c],'bat usage'])
        #print(pm_t1_bowl,runs_t2[c]-game_avg[c],pm_t2_bowl,runs_t1[c]-game_avg[c])
        final.loc[final['team']==home[c],'bowl +/-'] = final.loc[final['team']==home[c],'bowl +/-'] + ((runs_t2[c]-game_avg[c]-pm_t1_bowl) * final.loc[final['team']==home[c],'bowl usage'])
        final.loc[final['team']==opps[c],'bowl +/-'] = final.loc[final['team']==opps[c],'bowl +/-'] + ((runs_t1[c]-game_avg[c]-pm_t2_bowl) * final.loc[final['team']==opps[c],'bowl usage'])
        c+=1
    final['+/-'] = (final['bat +/-']) - (final['bowl +/-'])
    final['+/-'] = round(final['+/-'],2)
    #print(final.loc[final['team']==home[0],'bat +/-'].sum())
    #print(final.loc[final['team']==opps[0],'bat +/-'].sum())
    #print(final.loc[final['team']==home[0],'bowl +/-'].sum())
    #print(final.loc[final['team']==opps[0],'bowl +/-'].sum())
    final.drop(['bat +/-','bowl +/-'], axis=1, inplace=True)
    return final

def execute_projections():
    c = 0;
    for x in home:
        a = home[c]
        b = opps[c]
        v = venue[c]
        try: tbf = team_bat_first[c]
        except IndexError: tbf = ''
        if (c==0):
            if(coversoff == 1): (bat_game,bowl_game,field_game,table,runs_t1,runs_t2) = coversoff_projection(a,b,input_file1,input_file,f,v,tbf,reduced_length)
            elif(cricdraft == 1): (bat_game,bowl_game,field_game,table,runs_t1,runs_t2) = cricdraft_projection(a,b,input_file1,input_file,f,v,tbf,reduced_length)
            elif(ex22 == 1): (bat_game,bowl_game,field_game,table,runs_t1,runs_t2) = ex22_projection(a,b,input_file1,input_file,f,v,tbf,reduced_length)
            else: (bat_game,bowl_game,field_game,table,runs_t1,runs_t2) = gw_projection(a,b,input_file1,input_file,f,v,tbf,reduced_length)
            runs_t1 = [runs_t1]
            runs_t2 = [runs_t2]
        else : 
            if(coversoff == 1): (bat,bowl,field,table,r_t1,r_t2) = coversoff_projection(a,b,input_file1,input_file,f,v,tbf,reduced_length)
            elif(cricdraft == 1): (bat,bowl,field,table,r_t1,r_t2) = cricdraft_projection(a,b,input_file1,input_file,f,v,tbf,reduced_length)
            elif(ex22 == 1): (bat,bowl,field,table,r_t1,r_t2) = ex22_projection(a,b,input_file1,input_file,f,v,tbf,reduced_length)
            else: (bat,bowl,field,table,r_t1,r_t2) = gw_projection(a,b,input_file1,input_file,f,v,tbf,reduced_length)
            bat_game = pd.concat([bat_game,bat])
            bowl_game = pd.concat([bowl_game,bowl])
            field_game = pd.concat([field_game,field])
            runs_t1.append(r_t1)
            runs_t2.append(r_t2)
        c = c + 1
    a_final = final_projections(bat_game,bowl_game,field_game,runs_t1,runs_t2)
    return a_final,table,bat_game,bowl_game,field_game

a_final,a_table,bat_game,bowl_game,field_game = execute_projections()
#a_final.rename(columns = {'bat usage':'balls faced', 'bowl usage':'overs bowled'}, inplace = True)
#a_final['balls faced'] = a_final['balls faced']*120*f*reduced_length
#a_final['overs bowled'] = a_final['overs bowled']*20*f*reduced_length
#a_final['overs bowled'] = (a_final['overs bowled'] - (a_final['overs bowled'] % 1)) + ((a_final['overs bowled'] % 1)*0.6)
bat_game['runs']=(bat_game['usage']*bat_game['runs/ball']*f*120*reduced_length)
bowl_game['wickets']=(bowl_game['usage']*bowl_game['wickets/ball']*f*120*reduced_length)

#%% create a distribution of players
#players classified as wicketkeepers by dream 11
d11_keeper_eligibility = ['SV Samson','H Klaasen','RD Rickelton','JM Sharma','RR Pant','DP Conway','TA Blundell','BKG Mendis',
                          'S Samarawickrama','MJ Hay','JP Inglis','Mohammad Rizwan','Usman Khan','Haseebullah Khan','YH Bhatia',
                          'N Faltum','JC Buttler','PD Salt','MS Pepper','SD Hope','N Pooran','KNM Fernando','AT Carey','KL Rahul',
                          'Dhruv Jurel','G Redmayne','S Reid','L Lee','SJ Bryce','L Winfield-Hill','J Gumbie','T Marumani']

def solver(f_points):
    duplicate = f_points.copy()
    dummy_team = duplicate['team'][0]
    duplicate['dummy_team'] = np.where(duplicate['team']==dummy_team,1,0)
    duplicate['WK'] = np.where(duplicate['Pos']==1,1,0)
    duplicate['BAT'] = np.where(duplicate['Pos']==2,1,0)
    duplicate['ALL'] = np.where(duplicate['Pos']==3,1,0)
    duplicate['BOWL'] = np.where(duplicate['Pos']==4,1,0)
    duplicate['UC'] = np.where(duplicate['Pos']==5,1,0)
    duplicate['Sel'] = 0
    model = LpProblem(name="resource-allocation", sense=LpMaximize)
    
    # Define the decision variables
    x = {i: LpVariable(name=f"x{i}", cat="Binary") for i in range (0, len(duplicate))}
    c = {i: LpVariable(name=f"c{i}", cat="Binary") for i in range (0, len(duplicate))}
    vc = {i: LpVariable(name=f"vc{i}", cat="Binary") for i in range (0, len(duplicate))}
    
    # Add constraints
    model += (lpSum(x.values()) == 11, "total players in team")
    model += (lpSum(c.values()) == 1, "captain")
    model += (lpSum(vc.values()) == 1, "vice captain")
    model += (lpSum( x[k] * duplicate['cost'][k] for k in range (0, len(duplicate))) <= 100, "cost")
    model += (lpSum( x[k] * duplicate['WK'][k] for k in range (0, len(duplicate))) >= 1, "WK")
    #model += (lpSum( x[k] * duplicate['BAT'][k] for k in range (0, len(duplicate))) >= 1, "BAT")
    #model += (lpSum( x[k] * duplicate['ALL'][k] for k in range (0, len(duplicate))) >= 1, "ALL")
    #model += (lpSum( x[k] * duplicate['BOWL'][k] for k in range (0, len(duplicate))) >= 1, "BOWL")
    model += (lpSum( x[k] * duplicate['UC'][k] for k in range (0, len(duplicate))) >= 1, "UC")
    model += (lpSum( x[k] * duplicate['dummy_team'][k] for k in range (0, len(duplicate))) >= 1, "Team min")
    model += (lpSum( x[k] * duplicate['dummy_team'][k] for k in range (0, len(duplicate))) <= 10, "Team max")
    
    for k in range (0, len(duplicate)): model += (x[k]-c[k]-vc[k] >= 0,f"unique C-VC {k}")
    
    # Set objective
    model += lpSum( (x[k]+c[k]+0.5*vc[k]) * duplicate['total'][k] for k in range (0, len(duplicate)))
    
    # Solve the optimization problem
    status = model.solve(solver=GLPK(msg=False))    
    #print(f"{LpStatus[model.status]}, xPts {model.objective.value()}")
    xpts = model.objective.value()
    
    #for var in model.variables(): print(var.name,var.value())
    for name, constraint in model.constraints.items():
        if(name == 'cost'): cost = 100+constraint.value()
        #print(f"{name}: {constraint.value()}")   
    for k in range (0, len(duplicate)): duplicate['Sel'][k] = x[k].value() + c[k].value() + 0.5*vc[k].value()
    
    return duplicate,xpts,cost

def iterator(f_points,n,f):
    f_points.reset_index(drop=True, inplace=True)
    a_team = [["1","2","3","4","5","6","7","8",'9','10','11',"C","VC","xPts",'Cost']];
    k=1
    while(k<n+1):      
        #random noise introduced in mins/efficiency
        f_points_copy = f_points.copy()
        if(k>1):
            XP = get_truncated_normal(mean=f_points['total'],sd=f_points['total'], low=0, upp=300)
            f_points_copy['total'] = XP.rvs()
            #print(f_points_copy[['player','total']])
        solution,xPts,cost = solver(f_points_copy)
        solution = solution.sort_values(by=['batting order'],ascending=True)
        names = solution.loc[solution['Sel'] >= 1, 'player']
        names = names.to_list()
        
        xPts=0; #xmins=0
        for j in names:
            #xmins += f_points.loc[f_points['Name'] == j, 'MIN'].values[0]
            xPts += f_points.loc[f_points['player'] == j, 'total'].values[0]
        cap = solution.loc[solution['Sel'] == 2, 'player']
        cap = cap.to_list()
        vice = solution.loc[solution['Sel'] == 1.5, 'player']
        vice = vice.to_list()
        #print(solution)
        xPts += f_points.loc[f_points['player'] == cap[0], 'total'].values[0] + (f_points.loc[f_points['player'] == vice[0], 'total'].values[0])/2
        #xmins = round(xmins,2)
        names = names + cap + vice + [xPts] + [cost]
        if(xPts in chain(*a_team)): k = k - 1
        else: a_team.append(names)
        k = k + 1   
    return a_team

def maximizer(f_points,sims):
    f_points.reset_index(drop=True, inplace=True)
    a_leader = f_points[['player','team','total']]
    a_leader['iterations'] = 0
    k=0
    while(k<sims):      
        #random noise introduced in mins/efficiency
        f_points_copy = f_points.copy()
        if(k>0):
            XP = get_truncated_normal(mean=f_points['total'],sd=f_points['total'], low=0, upp=300)
            f_points_copy['total'] = XP.rvs()
            
        max_guy = f_points_copy.loc[f_points_copy['total'] == max(f_points_copy['total']),'player'].values[0]
        a_leader.loc[a_leader['player']==max_guy,'iterations'] += 1
        k+=1
    a_leader['iterations'] = a_leader['iterations']/sims
    a_leader['break even'] = 0.75/a_leader['iterations']
    return a_leader
    
if(gw == 0):
    a_final['cost'] = 1.0
    a_final['Pos'] = 5
    a_final.loc[a_final['player'].isin(d11_keeper_eligibility), 'Pos'] = 1
    a_team = iterator(a_final,unique_combos,f)
    a_team = pd.DataFrame(a_team)
    a_team.columns = a_team.iloc[0];a_team = a_team.drop(0)
    a_team = a_team.apply(pd.to_numeric, errors='ignore')
    a_team.drop('Cost', axis=1, inplace=True)
    a_team = a_team.sort_values(['xPts'], ascending=[False])
    a_final.drop('cost', axis=1, inplace=True)
    a_final.drop('Pos', axis=1, inplace=True)
    
a_final = a_final.sort_values(['total'], ascending=[False])

if(ex22 == 1):
    a_leader = maximizer(a_final,10000)

#%% dump projections to the desired file
date = now.strftime("%m-%d-%y")
def write_to_excel(date,output_dump,final):
    try:
        book = load_workbook(output_dump)
        with pd.ExcelWriter(output_dump, engine = 'openpyxl',mode='a', if_sheet_exists = 'replace') as writer:
            writer.book = book
            writer.sheets = {ws.title:ws for ws in book.worksheets}
            final.to_excel(writer, sheet_name=f'{date}',index=False)
    except FileNotFoundError:
        with pd.ExcelWriter(output_dump) as writer:        
            final.to_excel(writer, sheet_name=f'{date}',index=False)
            
def cricdraft_write(gw,output_dump,final):
    try:
        book = load_workbook(output_dump)
        with pd.ExcelWriter(output_dump, engine = 'openpyxl',mode='a', if_sheet_exists = 'replace') as writer:
            writer.book = book
            writer.sheets = {ws.title:ws for ws in book.worksheets}
            final.to_excel(writer, sheet_name = f'gw{gw}',index=False)
    except FileNotFoundError:
        with pd.ExcelWriter(output_dump) as writer:        
            final.to_excel(writer, sheet_name=f'gw{gw}', index=False)

if(write != 0 and gw == 0):
    write_to_excel(date,output_dump,a_final)
elif(coversoff == 0 and ex22 == 0 and cricdraft ==0 and write == 1):
    output_dump = f'{path}/outputs/{comp}_gw{gw}_{year}.xlsx'
    with pd.ExcelWriter(output_dump) as writer:        
        a_final.to_excel(writer, sheet_name='Points',index=False)
elif(coversoff == 1 and gw > 0 and write == 1):
    output_dump = f'{path}/outputs/{comp}_cricbattle_{year}.xlsx'
    with pd.ExcelWriter(output_dump) as writer:        
        a_final.to_excel(writer, sheet_name="Points", index=False)
elif(cricdraft == 1 and gw > 0 and write == 1):
    output_dump = f'{path}/outputs/{comp}_draft_{year}.xlsx'
    cricdraft_write(gw,output_dump,a_final)
