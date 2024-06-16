# -*- coding: utf-8 -*-
"""
Created on Tue Apr 18 16:46:51 2023
Generating fantasy xPts for dream 11 based on projections
@author: Subramanya.Ganti
"""
#%% choose the teams which are playing
import pandas as pd
import numpy as np
from scipy.stats import truncnorm
import datetime as dt
from openpyxl import load_workbook
import random
from collections import Counter
pd.options.mode.chained_assignment = None  # default='warn'
np.seterr(all='ignore')

comp = 'blast'; year = '24'; unique_combos = 4
#if another league data is used as a proxy then set 1
home=[]; opps=[]; venue = []; proxy = 0; custom = 0
#date based game selection if 0, else specific gameweek or entire season
gw = 0; write = 0
#select teams manually
#home = ["Canada"]; opps = ["Australia"]; venue = ["Providence"]
#home = ["Worcestershire"]; opps = ["Lancashire"]; venue = ["Worcester"]
#select custom date
#custom = dt.datetime(2024,6,9) #year,month,date
#type of scoring system, default dream 11
coversoff = 0; ex22 = 0; cricdraft = 1

#frauds like ben stokes who bowl whenever they feel like it
not_bowling_list = ['BA Stokes']
#frauds who suddenly decide to bat in a different position
custom_position_list = [['DR Mousley',4],['SR Patel',7],['BA Carse',5],['BA Raine',7],['KS Carlson',1],['EJ Byrom',2],
                        ['D Elgar',1],['MS Pepper',3],['AM Rossington',2],['PI Walter',5],['DR Sams',7],
                        ['B Swanepoel',8],['XC Bartlett',11],['LWP Wells',2],['GJ Bell',3],['JL du Plooy',4],
                        ['DJ Willey',3],['RS Bopara',4],['LW James',4],['TB Abell',4],['TK Curran',6],['OJ Price',7],
                        ['HD Ward',1],['OJ Carter',4],['DJ Lamb',5],['FJ Hudson-Prentice',8],['GH Roderick',6],
                        ['AJ Hose',4],['JJ Cobb',3],['MJ Waite',1],['EA Brookes',5],['A Lyth',1],['James Bracey',3],
                        ['JJ Bohannon',1],['CJ Anderson',6]]

#%% find projections for the games in question
path = 'C:/Users/Subramanya.Ganti/Downloads/cricket'
if(proxy == 1): input_file = f"{path}/projections/{comp}_proxy_projections.xlsx"
else: input_file = f"{path}/projections/{comp}_projections.xlsx"
input_file1 = f"{path}/summary/{comp}_summary.xlsx"
output_dump = f"{path}/outputs/{comp}_{year}_game.xlsx"
game_avg = []
home_win = []

custom_position_list = pd.DataFrame(custom_position_list, columns=['player', 'position'])
custom_position_list = custom_position_list.sort_values(by=['position'],ascending=True)
with pd.ExcelWriter(f'{path}/custom_positions.xlsx') as writer:
    custom_position_list.to_excel(writer, sheet_name="custom positions", index=False)

def fixtures_file(now,comp,gw):
    fixtures = f"{path}/schedule.xlsx"
    if(comp == 'hundredw'): fixtures = pd.read_excel(fixtures,f'hundred {year}')
    else: fixtures = pd.read_excel(fixtures,f'{comp} {year}')
    temp = dt.datetime(1899, 12, 30)    # Note, not 31st Dec but 30th!
    delta = now - temp
    now = float(delta.days)
    home = []; opps = []

    for x in fixtures.values:
        if(gw == 0 and x[1] == now):
            home.append(x[3])
            opps.append(x[4])
            venue.append(x[5])
        elif(gw == x[2]):
            home.append(x[3])
            opps.append(x[4])
            venue.append(x[5])
            
    return (home,opps)

if(home==[] and opps==[]):
    if(custom !=0): now = custom
    else: now = dt.datetime.now()
    (home,opps) = fixtures_file(now,comp,gw)
else:
    now = dt.datetime.now()

if(comp=='hundred' or comp=='hundredw'):
    f = (5/6); #hundred
elif(comp=='odi' or comp=='odiw' or comp=='odiq' or comp=='rlc'):
    f = 2.5;   #odi
elif(comp=='tests' or comp == 'cc' or comp == 'shield' or comp == 'testsw'):
    f = 11.25; #test
else:
    f = 1;     #assume its a t20 by default

from usage import *
def adj(x,f,bat):
    if(bat == 1):
        #print(x)
        x[6] = x[6]*f
        x[5] = x[5]*f
    else:
        x[4] = x[4]*f
        x[11] = x[11]*f
        
    x[14] = x[14]*f
    x[15] = x[15]*f
    x[16] = x[16]*f
    x[17] = x[17]*f
    x[13] = x[13]*f
    x[12] = 1 - (x[14]+x[15]+x[16]+x[17]+x[13])
    return x

def get_truncated_normal(mean=0, sd=1, low=0, upp=10):
    return truncnorm((low - mean) / sd, (upp - mean) / sd, loc=mean, scale=sd)

def adj_bowl_usage(t1,t2,bowl_game):
    #print(bowl_game.loc[bowl_game['team']==t1,'usage'].sum())
    #print(bowl_game.loc[bowl_game['team']==t2,'usage'].sum())
    t1_bowl = bowl_game.loc[bowl_game['team']==t1,'usage'].sum()
    t2_bowl = bowl_game.loc[bowl_game['team']==t2,'usage'].sum()
    delta = 1
    while(delta > 0.00000001):
        bowl_game['usage'] = np.minimum(bowl_game['usage'],0.2)
        delta = t1_bowl - bowl_game.loc[bowl_game['team']==t1,'usage'].sum()
        count_t1 = len(bowl_game.loc[bowl_game['team']==t1,'team'])
        bowl_game.loc[bowl_game['team']==t1,'usage'] = bowl_game.loc[bowl_game['team']==t1,'usage'] + (delta/count_t1)
        #print(delta)
    delta = 1
    while(delta > 0.00000001):
        bowl_game['usage'] = np.minimum(bowl_game['usage'],0.2)
        delta = t2_bowl - bowl_game.loc[bowl_game['team']==t2,'usage'].sum()
        count_t2 = len(bowl_game.loc[bowl_game['team']==t2,'team'])
        bowl_game.loc[bowl_game['team']==t2,'usage'] = bowl_game.loc[bowl_game['team']==t2,'usage'] + (delta/count_t2)
        #print(delta)        
    return bowl_game

def game_results(bat_game,factor,t1,t2,s1,s2,c1,c2,league_avg,summary):    
    if(factor == 5/6): factor = 1
    runs_t1 = round(s1*c2*league_avg*(bat_game.loc[bat_game['team']==t1,'usage'].sum()),2)
    runs_t2 = round(s2*c1*league_avg*(bat_game.loc[bat_game['team']==t2,'usage'].sum()),2)
    win_t1 = round(pow(runs_t1,15.5/factor)/(pow(runs_t1,15.5/factor)+pow(runs_t2,15.5/factor)),3)
    win_t2 = round(pow(runs_t2,15.5/factor)/(pow(runs_t1,15.5/factor)+pow(runs_t2,15.5/factor)),3)
    if(factor<11.25):
        print(t1,runs_t1,"runs in",round(20*factor*bat_game.loc[bat_game['team']==t1,'usage'].sum(),2),"overs")
        print(t2,runs_t2,"runs in",round(20*factor*bat_game.loc[bat_game['team']==t2,'usage'].sum(),2),"overs")
        print("P(win) :-",t1,win_t1,"-",win_t2,t2)
        global home_win
        home_win.append(win_t1)
    else:
        print(t1,"1st",round(runs_t1*0.6,2),"runs in",round(0.6*20*factor*bat_game.loc[bat_game['team']==t1,'usage'].sum(),2),"overs")
        print(t2,"1st",round(runs_t2*0.6,2),"runs in",round(0.6*20*factor*bat_game.loc[bat_game['team']==t2,'usage'].sum(),2),"overs")
        print(t1,"2nd",round(runs_t1*0.4,2),"runs in",round(0.4*20*factor*bat_game.loc[bat_game['team']==t1,'usage'].sum(),2),"overs")
        print(t2,"2nd",round(runs_t2*0.4,2),"runs in",round(0.4*20*factor*bat_game.loc[bat_game['team']==t2,'usage'].sum(),2),"overs")
    print(" ")
    return (runs_t1,runs_t2)

def cricdraft_bonus(SR,bf,rs,ECON,wkts,bb):
    c = 0; bat_total = 0; bowl_total = 0
    while(c<1000):
        bfi = bf.rvs()
        SRi = SR.rvs()
        rsi = bfi*SRi/100
        milstonei = np.where((rsi>=50)&(rsi<100), 25, 0) + np.where((rsi>=100)&(rsi<75), 50, 0) + np.where(rsi>=150, 75, 0)
        SR_bonusi = np.where((SRi>=0)&(SRi<50), -20, 0) + np.where((SRi>=50)&(SRi<70), -15, 0) + np.where((SRi>=70)&(SRi<90), -10, 0) + \
                    np.where((SRi>=90)&(SRi<100), -5, 0) + np.where((SRi>=140)&(SRi<170), 0.5*rsi, 0) + np.where((SRi>=170)&(SRi<200), rsi, 0) + \
                    np.where((SRi>=200)&(SRi<230), 1.5*rsi, 0) + np.where(SRi>=230, 2*rsi, 0)
        qualifyi =  np.where((rsi>=16)|(bfi>=12), 1, 0)
        bat_total += milstonei + qualifyi*SR_bonusi
        bbi = bb.rvs()
        ECONi = ECON.rvs()
        rci = bbi*ECONi/6
        wktsi = wkts.rvs()
        wickets_multiplei = np.where((wktsi>=3)&(wktsi<4), 25, 0) + np.where((wktsi>=4)&(wktsi<5), 35, 0) + np.where((wktsi>=5)&(wktsi<6), 50, 0) + \
                            np.where((wktsi>=6)&(wktsi<7), 75, 0) + np.where(wktsi>=7, 100, 0)
        ECON_bonusi = np.where((ECONi>=0)&(ECONi<4), 100, 0) + np.where((ECONi>=4)&(ECONi<5), 50, 0) + np.where((ECONi>=5)&(ECONi<6), 25, 0) + np.where((ECONi>=6)&(ECONi<7), 15, 0) + \
                      np.where((ECONi>=9)&(ECONi<10), -5, 0) +np.where((ECONi>=10)&(ECONi<12), -10, 0) + np.where((ECONi>=12)&(ECONi<14), -15, 0) + np.where(ECONi>=14, -20, 0)
        qualify2i = np.where((rci>=16)|(bbi>=12), 1, 0)
        bowl_total += wickets_multiplei + qualify2i*ECON_bonusi
        c += 1
    #print (bat_total/1000)
    #print (bowl_total/1000)
    return ((bat_total/1000),(bowl_total/1000))

def base_calculations(a,b,input_file1,input_file,factor,v):
    t1 = a
    t2 = b
    print(t1,"vs",t2)
    print('Venue -',v)
    venues = pd.read_excel(input_file,'Venue factors')
    vrf = (venues.loc[venues['Venue']==v,'runs'].sum())**0.5        #square root because the factor is for the game as a whole
    vwf = (venues.loc[venues['Venue']==v,'wkts'].sum())**0.5
    if(vrf==0):vrf=1
    if(vwf==0):vwf=1
    (summary,bat,bowl) = h2h_alt(input_file1,input_file,1,factor)
    bowl = adj_bowl_usage(t1,t2,bowl)
    summary = summary.apply(pd.to_numeric, errors='ignore')
    league_avg = (summary.loc[(summary['Team']!="Free Agent"),'runs bat'].mean() + summary.loc[(summary['Team']!="Free Agent"),'runs bowl'].mean())/2
    bat['runs/ball'] = bat['runs/ball']*vrf
    bat['SR'] = bat['runs/ball']*100
    bowl['runs/ball'] = bowl['runs/ball']*vrf
    bat['xSR'] = bat['xSR']*vrf
    bowl['xECON'] = bowl['xECON']*vrf
    bowl = bowl[~bowl['bowler'].isin(not_bowling_list)]
    
    if(factor<0.75):
        #old logic for balls faced
        ut1 = (bat.loc[bat['team']==t2,'usage'].sum())/(bowl.loc[bowl['team']==t1,'usage'].sum())
        ut2 = (bat.loc[bat['team']==t1,'usage'].sum())/(bowl.loc[bowl['team']==t2,'usage'].sum())
        bowl.loc[bowl['team']==t1,'usage'] = bowl.loc[bowl['team']==t1,'usage'] * ut1
        bowl.loc[bowl['team']==t2,'usage'] = bowl.loc[bowl['team']==t2,'usage'] * ut2
    else:
        #print(t1,"bowling",bowl.loc[bowl['team']==t1,'usage'].sum()*120*factor)
        #print(t2,"bowling",bowl.loc[bowl['team']==t2,'usage'].sum()*120*factor)
        #print(t1,"batting",bat.loc[bat['team']==t1,'usage'].sum()*120*factor)
        #print(t2,"batting",bat.loc[bat['team']==t2,'usage'].sum()*120*factor)
        #print(t1,"bowling wickets",(bowl.loc[bowl['team']==t1,'usage']*bowl.loc[bowl['team']==t1,'wickets/ball']).sum()*120*factor)
        #print(t2,"bowling wickets",(bowl.loc[bowl['team']==t2,'usage']*bowl.loc[bowl['team']==t2,'wickets/ball']).sum()*120*factor)
        #print(t1,"batting wickets",(bat.loc[bat['team']==t1,'usage']*bat.loc[bat['team']==t1,'wickets/ball']).sum()*120*factor)
        #print(t2,"batting wickets",(bat.loc[bat['team']==t2,'usage']*bat.loc[bat['team']==t2,'wickets/ball']).sum()*120*factor)
        
        #new logic for balls faced
        bb_t1 = bowl.loc[bowl['team']==t1,'usage'].sum()
        bb_t2 = bowl.loc[bowl['team']==t2,'usage'].sum()
        bf_t1 = bat.loc[bat['team']==t1,'usage'].sum()
        bf_t2 = bat.loc[bat['team']==t2,'usage'].sum()
        wbo_t1 = (bowl.loc[bowl['team']==t1,'usage']*bowl.loc[bowl['team']==t1,'wickets/ball']).sum()
        wbo_t2 = (bowl.loc[bowl['team']==t2,'usage']*bowl.loc[bowl['team']==t2,'wickets/ball']).sum()
        wba_t1 = (bat.loc[bat['team']==t1,'usage']*bat.loc[bat['team']==t1,'wickets/ball']).sum()
        wba_t2 = (bat.loc[bat['team']==t2,'usage']*bat.loc[bat['team']==t2,'wickets/ball']).sum()
        
        ut3 = min(pow(bb_t1*bf_t2*wba_t2/wbo_t1,0.5),1)
        ut4 = min(pow(bb_t2*bf_t1*wba_t1/wbo_t2,0.5),1)
        ut1b = ut3 / (bowl.loc[bowl['team']==t1,'usage'].sum())
        ut2b = ut4 / (bowl.loc[bowl['team']==t2,'usage'].sum())
        ut1bat = ut4 / (bat.loc[bat['team']==t1,'usage'].sum())
        ut2bat = ut3 / (bat.loc[bat['team']==t2,'usage'].sum())
        bowl.loc[bowl['team']==t1,'usage'] = bowl.loc[bowl['team']==t1,'usage'] * ut1b
        bowl.loc[bowl['team']==t2,'usage'] = bowl.loc[bowl['team']==t2,'usage'] * ut2b
        bat.loc[bat['team']==t2,'usage'] = bat.loc[bat['team']==t2,'usage'] * ut2bat
        bat.loc[bat['team']==t1,'usage'] = bat.loc[bat['team']==t1,'usage'] * ut1bat
        #print(bowl.loc[bowl['team']==t1,'usage'].sum(),bowl.loc[bowl['team']==t2,'usage'].sum(),ut3,ut4)
    
    bowl_t2 = (bowl.loc[bowl['team']==t2,'usage']*bowl.loc[bowl['team']==t2,'runs/ball']*factor*120).sum()
    bowl_t1 = (bowl.loc[bowl['team']==t1,'usage']*bowl.loc[bowl['team']==t1,'runs/ball']*factor*120).sum()
    bat_t2 = (bat.loc[bat['team']==t2,'usage']*bat.loc[bat['team']==t2,'runs/ball']*factor*120).sum() + (bowl.loc[bowl['team']==t1,'usage']*bowl.loc[bowl['team']==t1,'extras/ball']*factor*120).sum()
    bat_t1 = (bat.loc[bat['team']==t1,'usage']*bat.loc[bat['team']==t1,'runs/ball']*factor*120).sum() + (bowl.loc[bowl['team']==t2,'usage']*bowl.loc[bowl['team']==t2,'extras/ball']*factor*120).sum()
    #print("bowl t2",bowl_t2);
    #print("bowl t1",bowl_t1)
    #print("bat t2",bat_t2)
    #print("bat t1",bat_t1)
    #print(league_avg)
    s1 = bowl_t2/(league_avg*bowl.loc[bowl['team']==t2,'usage'].sum())
    c1 = bat_t2/(league_avg*bat.loc[bat['team']==t2,'usage'].sum())
    s2 = bowl_t1/(league_avg*bowl.loc[bowl['team']==t1,'usage'].sum())
    c2 = bat_t1/(league_avg*bat.loc[bat['team']==t1,'usage'].sum())
    
    w1 = 0; w2 = 0;
    w1_bowl = 0; w2_bowl = 0;

    #bat = pd.read_excel(input_file,'MDist bat')
    bat = bat.drop(['RSAA'],axis=1)
    bat = bat.drop(['OAA'],axis=1)
    bat = bat.drop(['xballs/wkt'],axis=1)
    #bat = bat.drop(['xSR'],axis=1)
    
    #bowl = pd.read_excel(input_file,'MDist bowl')
    bowl = bowl.drop(['RCAA'],axis=1)
    bowl = bowl.drop(['WTAA'],axis=1)
    #bowl = bowl.drop(['xECON'],axis=1)
    bowl = bowl.drop(['xSR'],axis=1)
    bat_game = [["batsman","season","team","usage","balls/wkt","SR","runs/ball","wickets/ball","pp usage","mid usage","setup usage","death usage","dots/ball","1s/ball","2s/ball","3s/ball","4s/ball","6s/ball","xPts","xSR"]]
    bowl_game = [["bowler","season","team","usage","ECON","SR","wickets/ball","pp usage","mid usage","setup usage","death usage","runs/ball","dots/ball","1s/ball","2s/ball","3s/ball","4s/ball","6s/ball","extras/ball","xPts","xECON"]]

    w_avg_bat = bat.loc[(bat['team']!="Free Agent"),'usage']*bat.loc[(bat['team']!="Free Agent"),'wickets/ball']
    w_avg_bowl = bowl.loc[(bowl['team']!="Free Agent"),'usage']*bowl.loc[(bowl['team']!="Free Agent"),'wickets/ball']*(bat.loc[(bat['team']!="Free Agent"),'usage'].sum()/bowl.loc[(bowl['team']!="Free Agent"),'usage'].sum())
    w_avg_bat = w_avg_bat.sum()*120*factor/(summary.shape[0]-1)
    w_avg_bowl = w_avg_bowl.sum()*120*factor/(summary.shape[0]-1)
    w_avg = (w_avg_bat + w_avg_bowl)/2
    #print(w_avg_bat,w_avg_bowl,w_avg)
    #print(bat.columns)
    #print(bowl.columns)

    for x in bat.values:
        if(x[2] == t1):
            x = adj(x,s1,1)
            bat_game.append(x.tolist())
            w1 = w1 + x[3]*x[7]*120*factor
        if(x[2] == t2):
            x = adj(x,s2,1)
            bat_game.append(x.tolist())
            w2 = w2 + x[3]*x[7]*120*factor

    for x in bowl.values:
        if(x[2] == t1):
            x = adj(x,c1,0)
            bowl_game.append(x.tolist())
            w1_bowl = w1_bowl + x[3]*x[6]*120*factor
        if(x[2] == t2):
            x = adj(x,c2,0)
            bowl_game.append(x.tolist())
            w2_bowl = w2_bowl + x[3]*x[6]*120*factor

    bat_game = pd.DataFrame(bat_game)
    bat_game.columns = bat_game.iloc[0];bat_game = bat_game.drop(0)    
    bowl_game = pd.DataFrame(bowl_game)
    bowl_game.columns = bowl_game.iloc[0];bowl_game = bowl_game.drop(0)
    bat_game = bat_game.apply(pd.to_numeric, errors='ignore')
    bowl_game = bowl_game.apply(pd.to_numeric, errors='ignore')
    
    bat_game["xSR"] = bat_game["xPts"]
    bowl_game["xECON"] = bowl_game["xPts"]
    bat_game['wickets/ball'] = bat_game['wickets/ball']*vwf
    bowl_game['wickets/ball'] = bowl_game['wickets/ball']*vwf
    #hundred has only 5 balls per over
    if(factor==(5/6)): bowl_game['ECON'] = 5*bowl_game['runs/ball']; bowl_game['xECON'] = (5/6)*bowl_game['xECON']
    else: bowl_game['ECON'] = 6*bowl_game['runs/ball']
    
    #print(w1,w2,w1_bowl,w2_bowl,ut3,ut4)
    if(factor == 11.25):
        w1_bowl = min(w1_bowl/ut3,20)
        w2 = min(w2/ut3,20)
        w2_bowl = min(w2_bowl/ut4,20)
        w1 = min(w1/ut4,20)
        w_avg = w_avg * (bat_game['usage'].sum()/2)
        bat_game.loc[bat_game['team'] == t1, 'wickets/ball'] = bat_game.loc[bat_game['team'] == t1, 'wickets/ball']*(w2_bowl+0.7)/w_avg
        bat_game.loc[bat_game['team'] == t2, 'wickets/ball'] = bat_game.loc[bat_game['team'] == t2, 'wickets/ball']*(w1_bowl+0.7)/w_avg
        bat_game['balls/wkt'] = 1/bat_game['wickets/ball']
        bowl_game.loc[bowl_game['team'] == t1, 'wickets/ball'] = bowl_game.loc[bowl_game['team'] == t1, 'wickets/ball']*(w2-0.7)/w_avg
        bowl_game.loc[bowl_game['team'] == t2, 'wickets/ball'] = bowl_game.loc[bowl_game['team'] == t2, 'wickets/ball']*(w1-0.7)/w_avg
    else:
        w1_bowl = min(w1_bowl/ut3,10)
        w2 = min(w2/ut3,10)
        w2_bowl = min(w2_bowl/ut4,10)
        w1 = min(w1/ut4,10)
        bat_game.loc[bat_game['team'] == t1, 'wickets/ball'] = bat_game.loc[bat_game['team'] == t1, 'wickets/ball']*(w2_bowl+0.35)/w_avg
        bat_game.loc[bat_game['team'] == t2, 'wickets/ball'] = bat_game.loc[bat_game['team'] == t2, 'wickets/ball']*(w1_bowl+0.35)/w_avg
        bat_game['balls/wkt'] = 1/bat_game['wickets/ball']
        bowl_game.loc[bowl_game['team'] == t1, 'wickets/ball'] = bowl_game.loc[bowl_game['team'] == t1, 'wickets/ball']*(w2-0.35)/w_avg
        bowl_game.loc[bowl_game['team'] == t2, 'wickets/ball'] = bowl_game.loc[bowl_game['team'] == t2, 'wickets/ball']*(w1-0.35)/w_avg
        
    return (bat_game,bowl_game,summary,t1,t2,s1,s2,c1,c2,league_avg,w_avg)

print("")
def gw_projection(a,b,input_file1,input_file,factor,v):
    
    (bat_game,bowl_game,summary,t1,t2,s1,s2,c1,c2,league_avg,w_avg) = base_calculations(a,b,input_file1,input_file,factor,v)
    
    bat_game['xPts'] = 120*factor*bat_game['usage']*(bat_game['runs/ball']+bat_game['4s/ball']+2*bat_game['6s/ball'])
    bat_game['xPts'] = bat_game['xPts'] - 2*bat_game['wickets/ball']

    bat_game = bat_game.drop(['pp usage'],axis=1)
    bat_game = bat_game.drop(['mid usage'],axis=1)
    bat_game = bat_game.drop(['setup usage'],axis=1)
    bat_game = bat_game.drop(['death usage'],axis=1)
    bowl_game = bowl_game.drop(['pp usage'],axis=1)
    bowl_game = bowl_game.drop(['mid usage'],axis=1)
    bowl_game = bowl_game.drop(['setup usage'],axis=1)
    bowl_game = bowl_game.drop(['death usage'],axis=1)    
    
    SR = get_truncated_normal(mean=bat_game['SR'], sd=bat_game['SR']*0.36, low=0, upp=600)
    bf = get_truncated_normal(mean=bat_game['usage']*120*factor, sd=bat_game['usage']*120*0.6, low=0, upp=120*factor)
    rs = get_truncated_normal(mean=bat_game['runs/ball']*bat_game['usage']*120*factor, sd=bat_game['runs/ball']*bat_game['usage']*120*factor*0.6, low=0, upp=250*factor)
    ECON = get_truncated_normal(mean=bowl_game['ECON'], sd=bowl_game['ECON']*0.36, low=0, upp=36)
    wkts = get_truncated_normal(mean=bowl_game['wickets/ball']*bowl_game['usage']*120*factor, sd=bowl_game['wickets/ball']*bowl_game['usage']*120*factor*1.6, low=0, upp=10)
    
    if(factor != 11.25):
        bowl_game['usage'] = np.minimum(bowl_game['usage'],0.2)
        bowl_game['xPts'] = 120*factor*bowl_game['usage']*bowl_game['wickets/ball']*28
        
        if(factor == 1):    
            bat_game['xPts'] = bat_game['xPts'] + 6*(1-SR.cdf(170))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] + 4*(SR.cdf(170)-SR.cdf(150))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] + 2*(SR.cdf(150)-SR.cdf(130))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] - 2*(SR.cdf(70)-SR.cdf(60))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] - 4*(SR.cdf(60)-SR.cdf(50))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] - 6*(SR.cdf(50))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] + 4*(rs.cdf(50)-rs.cdf(30))
            bat_game['xPts'] = bat_game['xPts'] + 8*(rs.cdf(100)-rs.cdf(50))
            bat_game['xPts'] = bat_game['xPts'] + 16*(1-rs.cdf(100))
            bowl_game['xPts'] = bowl_game['xPts'] + 12*np.power(bowl_game['dots/ball'],6)*20*bowl_game['usage']*factor
            bowl_game['xPts'] = bowl_game['xPts'] + 6*(ECON.cdf(5))
            bowl_game['xPts'] = bowl_game['xPts'] + 4*(ECON.cdf(6)-ECON.cdf(5))
            bowl_game['xPts'] = bowl_game['xPts'] + 2*(ECON.cdf(7)-ECON.cdf(6))
            bowl_game['xPts'] = bowl_game['xPts'] - 2*(ECON.cdf(11)-ECON.cdf(10))
            bowl_game['xPts'] = bowl_game['xPts'] - 4*(ECON.cdf(12)-ECON.cdf(11))
            bowl_game['xPts'] = bowl_game['xPts'] - 6*(1-ECON.cdf(12))
            bowl_game['xPts'] = bowl_game['xPts'] + 4*(wkts.cdf(4)-wkts.cdf(3))
            bowl_game['xPts'] = bowl_game['xPts'] + 8*(wkts.cdf(5)-wkts.cdf(4))
            bowl_game['xPts'] = bowl_game['xPts'] + 16*(1-wkts.cdf(5))
            
            #legacy code
            #bat_game['xPts'] = bat_game['xPts'] + 6*(1-poisson.cdf(k=170,mu=bat_game['SR']))*(1-poisson.cdf(k=10,mu=120*bat_game['usage']*factor))
            #bat_game['xPts'] = bat_game['xPts'] + 4*(poisson.cdf(k=170,mu=bat_game['SR'])-poisson.cdf(k=150,mu=bat_game['SR']))*(1-poisson.cdf(k=10,mu=120*bat_game['usage']*factor))
            #bat_game['xPts'] = bat_game['xPts'] + 2*(poisson.cdf(k=150,mu=bat_game['SR'])-poisson.cdf(k=130,mu=bat_game['SR']))*(1-poisson.cdf(k=10,mu=120*bat_game['usage']*factor))
            #bat_game['xPts'] = bat_game['xPts'] - 2*(poisson.cdf(k=70,mu=bat_game['SR'])-poisson.cdf(k=60,mu=bat_game['SR']))*(1-poisson.cdf(k=10,mu=120*bat_game['usage']*factor))
            #bat_game['xPts'] = bat_game['xPts'] - 4*(poisson.cdf(k=60,mu=bat_game['SR'])-poisson.cdf(k=50,mu=bat_game['SR']))*(1-poisson.cdf(k=10,mu=120*bat_game['usage']*factor))
            #bat_game['xPts'] = bat_game['xPts'] - 6*poisson.cdf(k=50,mu=bat_game['SR'])*(1-poisson.cdf(k=10,mu=120*bat_game['usage']*factor))
            #bat_game['xPts'] = bat_game['xPts'] + 4*(poisson.cdf(k=50,mu=bat_game['runs/ball']*bat_game['usage']*factor*120)-poisson.cdf(k=30,mu=bat_game['runs/ball']*bat_game['usage']*factor*120))
            #bat_game['xPts'] = bat_game['xPts'] + 8*(poisson.cdf(k=100,mu=bat_game['runs/ball']*bat_game['usage']*factor*120)-poisson.cdf(k=50,mu=bat_game['runs/ball']*bat_game['usage']*factor*120))
            #bat_game['xPts'] = bat_game['xPts'] + 16*(1-poisson.cdf(k=100,mu=bat_game['runs/ball']*bat_game['usage']*factor*120))
            #bowl_game['xPts'] = bowl_game['xPts'] + 12*np.power(bowl_game['dots/ball'],6)*20*bowl_game['usage']*factor
            #bowl_game['xPts'] = bowl_game['xPts'] - 6*(1-poisson.cdf(k=12,mu=bowl_game['ECON']))
            #bowl_game['xPts'] = bowl_game['xPts'] - 4*(poisson.cdf(k=12,mu=bowl_game['ECON'])-poisson.cdf(k=11.,mu=bowl_game['ECON']))
            #bowl_game['xPts'] = bowl_game['xPts'] - 2*(poisson.cdf(k=11,mu=bowl_game['ECON'])-poisson.cdf(k=10,mu=bowl_game['ECON']))
            #bowl_game['xPts'] = bowl_game['xPts'] + 2*(poisson.cdf(k=7,mu=bowl_game['ECON'])-poisson.cdf(k=6,mu=bowl_game['ECON']))
            #bowl_game['xPts'] = bowl_game['xPts'] + 4*(poisson.cdf(k=6,mu=bowl_game['ECON'])-poisson.cdf(k=5,mu=bowl_game['ECON']))
            #bowl_game['xPts'] = bowl_game['xPts'] + 6*poisson.cdf(k=5,mu=bowl_game['ECON'])
            #bowl_game['xPts'] = bowl_game['xPts'] + 4*poisson.pmf(k=3,mu=bowl_game['wickets/ball']*bowl_game['usage']*120*factor)
            #bowl_game['xPts'] = bowl_game['xPts'] + 8*poisson.pmf(k=4,mu=bowl_game['wickets/ball']*bowl_game['usage']*120*factor)
            #bowl_game['xPts'] = bowl_game['xPts'] + 16*poisson.pmf(k=5,mu=bowl_game['wickets/ball']*bowl_game['usage']*120*factor)
            
        elif(factor == 2.5):
            bat_game['xPts'] = bat_game['xPts'] + 6*(1-SR.cdf(140))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] + 4*(SR.cdf(140)-SR.cdf(120))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] + 2*(SR.cdf(120)-SR.cdf(100))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] - 2*(SR.cdf(50)-SR.cdf(40))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] - 4*(SR.cdf(40)-SR.cdf(30))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] - 6*(SR.cdf(30))*(1-bf.cdf(10))
            bat_game['xPts'] = bat_game['xPts'] + 4*(rs.cdf(100)-rs.cdf(50))
            bat_game['xPts'] = bat_game['xPts'] + 8*(rs.cdf(200)-rs.cdf(100))
            bat_game['xPts'] = bat_game['xPts'] + 16*(1-rs.cdf(200))
            bowl_game['xPts'] = bowl_game['xPts'] + 12*np.power(bowl_game['dots/ball'],6)*20*bowl_game['usage']*factor
            bowl_game['xPts'] = bowl_game['xPts'] + 6*(ECON.cdf(2.5))
            bowl_game['xPts'] = bowl_game['xPts'] + 4*(ECON.cdf(3.5)-ECON.cdf(2.5))
            bowl_game['xPts'] = bowl_game['xPts'] + 2*(ECON.cdf(4.5)-ECON.cdf(3.5))
            bowl_game['xPts'] = bowl_game['xPts'] - 2*(ECON.cdf(8)-ECON.cdf(7))
            bowl_game['xPts'] = bowl_game['xPts'] - 4*(ECON.cdf(9)-ECON.cdf(8))
            bowl_game['xPts'] = bowl_game['xPts'] - 6*(1-ECON.cdf(9))
            bowl_game['xPts'] = bowl_game['xPts'] + 0*(wkts.cdf(4)-wkts.cdf(3))
            bowl_game['xPts'] = bowl_game['xPts'] + 4*(wkts.cdf(5)-wkts.cdf(4))
            bowl_game['xPts'] = bowl_game['xPts'] + 8*(1-wkts.cdf(5))
            
        elif(factor == 5/6):
            bat_game['xPts'] = bat_game['xPts'] + 5*(rs.cdf(50)-rs.cdf(30))
            bat_game['xPts'] = bat_game['xPts'] + 10*(rs.cdf(100)-rs.cdf(50))
            bat_game['xPts'] = bat_game['xPts'] + 20*(1-rs.cdf(100))
            bowl_game['xPts'] = bowl_game['xPts'] + 16*np.power(bowl_game['dots/ball'],6)*20*bowl_game['usage']*factor
            bowl_game['xPts'] = bowl_game['xPts'] + 5*(wkts.cdf(4)-wkts.cdf(3))
            bowl_game['xPts'] = bowl_game['xPts'] + 10*(wkts.cdf(5)-wkts.cdf(4))
            bowl_game['xPts'] = bowl_game['xPts'] + 20*(1-wkts.cdf(5))
            
    else:
        bowl_game['xPts'] = 120*factor*bowl_game['usage']*bowl_game['wickets/ball']*19        
        bat_game['xPts'] = bat_game['xPts'] + 4*(rs.cdf(100)-rs.cdf(50))
        bat_game['xPts'] = bat_game['xPts'] + 8*(rs.cdf(200)-rs.cdf(100))
        bat_game['xPts'] = bat_game['xPts'] + 16*(rs.cdf(300)-rs.cdf(200))
        bat_game['xPts'] = bat_game['xPts'] + 24*(1-rs.cdf(300))
        bowl_game['xPts'] = bowl_game['xPts'] + 4*(wkts.cdf(5)-wkts.cdf(4))
        bowl_game['xPts'] = bowl_game['xPts'] + 8*(1-wkts.cdf(5))
    
    runs_t1,runs_t2 = game_results(bat_game,factor,t1,t2,s1,s2,c1,c2,league_avg,summary)
    global game_avg
    game_avg.append((runs_t1+runs_t2)/2)
    return (bat_game,bowl_game,summary,runs_t1,runs_t2)

def cricdraft_projection(a,b,input_file1,input_file,factor,v):
    
    (bat_game,bowl_game,summary,t1,t2,s1,s2,c1,c2,league_avg,w_avg) = base_calculations(a,b,input_file1,input_file,factor,v)
    
    bat_game = bat_game.drop(['pp usage'],axis=1)
    bat_game = bat_game.drop(['mid usage'],axis=1)
    bat_game = bat_game.drop(['setup usage'],axis=1)
    bat_game = bat_game.drop(['death usage'],axis=1)
    bowl_game = bowl_game.drop(['pp usage'],axis=1)
    bowl_game = bowl_game.drop(['mid usage'],axis=1)
    bowl_game = bowl_game.drop(['setup usage'],axis=1)
    bowl_game = bowl_game.drop(['death usage'],axis=1)
    
    SR = get_truncated_normal(mean=bat_game['SR'], sd=bat_game['SR']*0.36, low=0, upp=600)
    bf = get_truncated_normal(mean=bat_game['usage']*120*factor, sd=bat_game['usage']*120*0.6, low=0, upp=120*factor)
    rs = get_truncated_normal(mean=bat_game['runs/ball']*bat_game['usage']*120*factor, sd=bat_game['runs/ball']*bat_game['usage']*120*factor*0.6, low=0, upp=250*factor)
    ECON = get_truncated_normal(mean=bowl_game['ECON'], sd=bowl_game['ECON']*0.36, low=0, upp=42)
    wkts = get_truncated_normal(mean=bowl_game['wickets/ball']*bowl_game['usage']*120*factor, sd=bowl_game['wickets/ball']*bowl_game['usage']*120*factor*1.6, low=0, upp=10)
    bb = get_truncated_normal(mean=bowl_game['usage']*120*factor, sd=bowl_game['usage']*120*0.2, low=0, upp=0.2*120*factor)
    
    if(factor != 11.25):
        bat_game['xPts'] = 120*factor*bat_game['usage']*(bat_game['runs/ball']+3*bat_game['6s/ball'])
        bowl_game['usage'] = np.minimum(bowl_game['usage'],0.2)
        bowl_game['xPts'] = 120*factor*bowl_game['usage']*bowl_game['wickets/ball']*30
        #bat_game['xPts'] = bat_game['xPts'] - 25*rs.cdf(1)*(1-bf.cdf(4))
        bat_bonus,bowl_bonus = cricdraft_bonus(SR,bf,rs,ECON,wkts,bb)
        if(factor <= 1):
            #multiplication of 2 gaussians is not a gaussian !
            bat_game['xPts'] += bat_bonus
            bowl_game['xPts'] += 25*np.power(bowl_game['dots/ball'],6)*20*bowl_game['usage']*factor
            bowl_game['xPts'] += bowl_bonus
            
        elif(factor == 2.5):
            bat_game['xPts'] = bat_game['xPts'] + 2.5*bat_game['runs/ball']*bat_game['usage']*120*factor*(1-SR.cdf(200))*(1-bf.cdf(24))
            bat_game['xPts'] = bat_game['xPts'] + 2*bat_game['runs/ball']*bat_game['usage']*120*factor*(SR.cdf(200)-SR.cdf(170))*(1-bf.cdf(24))
            bat_game['xPts'] = bat_game['xPts'] + 1.5*bat_game['runs/ball']*bat_game['usage']*120*factor*(SR.cdf(170)-SR.cdf(145))*(1-bf.cdf(24))
            bat_game['xPts'] = bat_game['xPts'] + bat_game['runs/ball']*bat_game['usage']*120*factor*(SR.cdf(145)-SR.cdf(120))*(1-bf.cdf(24))
            bat_game['xPts'] = bat_game['xPts'] + 0.5*bat_game['runs/ball']*bat_game['usage']*120*factor*(SR.cdf(120)-SR.cdf(95))*(1-bf.cdf(24))
            bat_game['xPts'] = bat_game['xPts'] - 5*(SR.cdf(80)-SR.cdf(65))*(1-bf.cdf(24))
            bat_game['xPts'] = bat_game['xPts'] - 10*(SR.cdf(65)-SR.cdf(50))*(1-bf.cdf(24))
            bat_game['xPts'] = bat_game['xPts'] - 15*(SR.cdf(50)-SR.cdf(40))*(1-bf.cdf(24))
            bat_game['xPts'] = bat_game['xPts'] - 20*(SR.cdf(40))*(1-bf.cdf(24))
            bat_game['xPts'] = bat_game['xPts'] + 25*(rs.cdf(100)-rs.cdf(50))
            bat_game['xPts'] = bat_game['xPts'] + 50*(rs.cdf(150)-rs.cdf(100))
            #bat_game['xPts'] = bat_game['xPts'] + 100*(1-rs.cdf(150))
            bowl_game['xPts'] = bowl_game['xPts'] + 15*np.power(bowl_game['dots/ball'],6)*20*bowl_game['usage']*factor
            bowl_game['xPts'] = bowl_game['xPts'] + 100*(ECON.cdf(2.5))*(1-bb.cdf(24))
            bowl_game['xPts'] = bowl_game['xPts'] + 70*(ECON.cdf(3.5)-ECON.cdf(2.5))*(1-bb.cdf(24))
            bowl_game['xPts'] = bowl_game['xPts'] + 45*(ECON.cdf(4.25)-ECON.cdf(3.5))*(1-bb.cdf(24))
            bowl_game['xPts'] = bowl_game['xPts'] + 25*(ECON.cdf(4.75)-ECON.cdf(4.25))*(1-bb.cdf(24))
            bowl_game['xPts'] = bowl_game['xPts'] + 10*(ECON.cdf(5.25)-ECON.cdf(4.75))*(1-bb.cdf(24))
            bowl_game['xPts'] = bowl_game['xPts'] - 5*(ECON.cdf(6.5)-ECON.cdf(5.75))*(1-bb.cdf(24))
            bowl_game['xPts'] = bowl_game['xPts'] - 10*(ECON.cdf(8)-ECON.cdf(6.5))*(1-bb.cdf(24))
            bowl_game['xPts'] = bowl_game['xPts'] - 15*(ECON.cdf(10.5)-ECON.cdf(8))*(1-bb.cdf(24))
            bowl_game['xPts'] = bowl_game['xPts'] - 20*(1-ECON.cdf(10.5))*(1-bb.cdf(24))
            bowl_game['xPts'] = bowl_game['xPts'] + 15*(wkts.cdf(4)-wkts.cdf(3))
            bowl_game['xPts'] = bowl_game['xPts'] + 25*(wkts.cdf(5)-wkts.cdf(4))
            bowl_game['xPts'] = bowl_game['xPts'] + 50*(1-wkts.cdf(5))
            #bowl_game['xPts'] = bowl_game['xPts'] + 75*(wkts.cdf(7)-wkts.cdf(6))
            #bowl_game['xPts'] = bowl_game['xPts'] + 100*(1-wkts.cdf(7))
            
    else:
        bowl_game['xPts'] = 120*factor*bowl_game['usage']*bowl_game['wickets/ball']*19        
        bat_game['xPts'] = bat_game['xPts'] + 4*(rs.cdf(100)-rs.cdf(50))
        bat_game['xPts'] = bat_game['xPts'] + 8*(rs.cdf(200)-rs.cdf(100))
        bat_game['xPts'] = bat_game['xPts'] + 16*(rs.cdf(300)-rs.cdf(200))
        bat_game['xPts'] = bat_game['xPts'] + 24*(1-rs.cdf(300))
        bowl_game['xPts'] = bowl_game['xPts'] + 4*(wkts.cdf(5)-wkts.cdf(4))
        bowl_game['xPts'] = bowl_game['xPts'] + 8*(1-wkts.cdf(5))
    
    runs_t1,runs_t2 = game_results(bat_game,factor,t1,t2,s1,s2,c1,c2,league_avg,summary)
    global game_avg
    game_avg.append((runs_t1+runs_t2)/2)
    return (bat_game,bowl_game,summary,runs_t1,runs_t2)

def ex22_projection(a,b,input_file1,input_file,factor,v):
    
    (bat_game,bowl_game,summary,t1,t2,s1,s2,c1,c2,league_avg,w_avg) = base_calculations(a,b,input_file1,input_file,factor,v)

    bat_game = bat_game.drop(['pp usage'],axis=1)
    bat_game = bat_game.drop(['mid usage'],axis=1)
    bat_game = bat_game.drop(['setup usage'],axis=1)
    bat_game = bat_game.drop(['death usage'],axis=1)
    bowl_game = bowl_game.drop(['pp usage'],axis=1)
    bowl_game = bowl_game.drop(['mid usage'],axis=1)
    bowl_game = bowl_game.drop(['setup usage'],axis=1)
    bowl_game = bowl_game.drop(['death usage'],axis=1)
    
    bat_game['xPts'] = 0.0; bowl_game['xPts'] = 0.0
    rs = get_truncated_normal(mean=bat_game['runs/ball']*bat_game['usage']*120*factor, sd=bat_game['runs/ball']*bat_game['usage']*120*factor*0.6, low=0, upp=250*factor)
    wkts = get_truncated_normal(mean=bowl_game['wickets/ball']*bowl_game['usage']*120*factor, sd=bowl_game['wickets/ball']*bowl_game['usage']*120*factor*1.6, low=0, upp=10)
    
    if(factor == 1):        
        bat_game['xPts'] = bat_game['xPts']+(bat_game['usage']*(bat_game['runs/ball']+bat_game['4s/ball']+2*bat_game['6s/ball'])*120*factor)
        bat_game['xPts'] += 7*(rs.cdf(100)-rs.cdf(50))
        bat_game['xPts'] += 15*(1-rs.cdf(100))
        bowl_game['usage'] = np.minimum(bowl_game['usage'],0.2)
        bowl_game['xPts'] += 20*bowl_game['usage']*bowl_game['wickets/ball']*120*factor
        bowl_game['xPts'] += 5*(wkts.cdf(4)-wkts.cdf(3))
        bowl_game['xPts'] += 8*(1-wkts.cdf(4))
        bowl_game['xPts'] += 7*np.power(bowl_game['dots/ball'],6)*20*bowl_game['usage']*factor
        
    elif(factor == 2.5):
        bat_game['xPts'] += bat_game['usage']*(bat_game['runs/ball']+bat_game['4s/ball']+2*bat_game['6s/ball'])*120*factor
        bat_game['xPts'] += 5*(rs.cdf(100)-rs.cdf(50))
        bat_game['xPts'] += 8*(1-rs.cdf(100))
        bowl_game['usage'] = np.minimum(bowl_game['usage'],0.2)
        bowl_game['xPts'] += 25*bowl_game['usage']*bowl_game['wickets/ball']*120*factor
        bowl_game['xPts'] += 5*(wkts.cdf(5)-wkts.cdf(4))
        bowl_game['xPts'] += 8*(1-wkts.cdf(5))
        bowl_game['xPts'] += 5*np.power(bowl_game['dots/ball'],6)*20*bowl_game['usage']*factor
        
    elif(factor == 5/6):         
        bat_game['xPts'] += bat_game['usage']*(bat_game['runs/ball']+bat_game['4s/ball']+2*bat_game['6s/ball'])*120*factor
        bat_game['xPts'] += 8*(1-rs.cdf(50))
        bowl_game['usage'] = np.minimum(bowl_game['usage'],0.2)
        bowl_game['xPts'] += 20*bowl_game['usage']*bowl_game['wickets/ball']*120*factor
        bowl_game['xPts'] += 5*(wkts.cdf(4)-wkts.cdf(3))
        bowl_game['xPts'] += 8*(1-wkts.cdf(4))
        
    else: #test match
        bat_game['xPts'] += bat_game['usage']*(0.5*bat_game['runs/ball']+bat_game['4s/ball']+3*bat_game['6s/ball'])*120*factor
        bat_game['xPts'] += 5*(rs.cdf(100)-rs.cdf(50))
        bat_game['xPts'] += 8*(rs.cdf(200)-rs.cdf(100))
        bat_game['xPts'] += 15*(1-rs.cdf(200))
        bowl_game['xPts'] += 12*bowl_game['usage']*bowl_game['wickets/ball']*120*factor
        bowl_game['xPts'] += 10*(1-wkts.cdf(5))
        bowl_game['xPts'] += 1*np.power(bowl_game['dots/ball'],6)*20*bowl_game['usage']*factor
     
    runs_t1,runs_t2 = game_results(bat_game,factor,t1,t2,s1,s2,c1,c2,league_avg,summary)
    global game_avg
    game_avg.append((runs_t1+runs_t2)/2)
    return (bat_game,bowl_game,summary,runs_t1,runs_t2)

def coversoff_projection(a,b,input_file1,input_file,factor,v):
    
    (bat_game,bowl_game,summary,t1,t2,s1,s2,c1,c2,league_avg,w_avg) = base_calculations(a,b,input_file1,input_file,factor,v)
        
    bat_game = bat_game.drop(['pp usage'],axis=1)
    bat_game = bat_game.drop(['mid usage'],axis=1)
    bat_game = bat_game.drop(['setup usage'],axis=1)
    bat_game = bat_game.drop(['death usage'],axis=1)
    #initialize points as 0
    bat_game['xPts'] = 0.0; bowl_game['xPts'] = 0.0
    SR = get_truncated_normal(mean=bat_game['SR'], sd=bat_game['SR']*0.36, low=0, upp=600)
    bf = get_truncated_normal(mean=bat_game['usage']*120*factor, sd=bat_game['usage']*120*0.6, low=0, upp=120*factor)
    rs = get_truncated_normal(mean=bat_game['runs/ball']*bat_game['usage']*120*factor, sd=bat_game['runs/ball']*bat_game['usage']*120*factor*0.6, low=0, upp=250*factor)
    ECON = get_truncated_normal(mean=bowl_game['ECON'], sd=bowl_game['ECON']*0.36, low=0, upp=36)
    wkts = get_truncated_normal(mean=bowl_game['wickets/ball']*bowl_game['usage']*120*factor, sd=bowl_game['wickets/ball']*bowl_game['usage']*120*factor*1.6, low=0, upp=10)
    bb = get_truncated_normal(mean=bowl_game['usage']*120*factor, sd=bowl_game['usage']*120*0.2, low=0, upp=0.2*120*factor)
    
    if(factor <= 1):
        #runs 4s and 6s have pts
        bat_game['xPts'] = 120*factor*bat_game['usage']*(bat_game['runs/ball']+10*bat_game['6s/ball']+6*bat_game['4s/ball'])
        #duck        
        bat_game['xPts'] = bat_game['xPts'] - 0*rs.cdf(1)*(1-bf.cdf(6))
        #less than x runs
        bat_game['xPts'] = bat_game['xPts'] - 0*(rs.cdf(5)-rs.cdf(1))*(1-bf.cdf(6))
        #run rate bonus
        #bat_game['xPts'] = bat_game['xPts'] + 120*factor*bat_game['usage']*bat_game['runs/ball'] - (4/3)*(bat_game['usage']*factor*120)
        #SR bonuses
        bat_game['xPts'] = bat_game['xPts'] + 120*(1-SR.cdf(200))*(1-rs.cdf(20))
        bat_game['xPts'] = bat_game['xPts'] + 100*(SR.cdf(200)-SR.cdf(180))*(1-rs.cdf(20))
        bat_game['xPts'] = bat_game['xPts'] + 80*(SR.cdf(180)-SR.cdf(160))*(1-rs.cdf(25))
        bat_game['xPts'] = bat_game['xPts'] + 60*(SR.cdf(160)-SR.cdf(140))*(1-rs.cdf(25))
        bat_game['xPts'] = bat_game['xPts'] - 80*(SR.cdf(130)-SR.cdf(100))*(1-bf.cdf(6))
        bat_game['xPts'] = bat_game['xPts'] - 100*(SR.cdf(100))*(1-bf.cdf(6))
        #milestone bonuses
        bat_game['xPts'] = bat_game['xPts'] + 20*(rs.cdf(35)-rs.cdf(20))
        bat_game['xPts'] = bat_game['xPts'] + 40*(rs.cdf(50)-rs.cdf(35))
        bat_game['xPts'] = bat_game['xPts'] + 70*(rs.cdf(65)-rs.cdf(50))
        bat_game['xPts'] = bat_game['xPts'] + 120*(rs.cdf(75)-rs.cdf(65))
        bat_game['xPts'] = bat_game['xPts'] + 150*(1-rs.cdf(75))
    elif(factor == 2.5):
        #runs 4s and 6s have pts
        bat_game['xPts'] = 120*factor*bat_game['usage']*(bat_game['runs/ball']+3*bat_game['6s/ball']+2*bat_game['4s/ball'])
        #duck        
        bat_game['xPts'] = bat_game['xPts'] - 20*rs.cdf(1)*(1-bf.cdf(30))
        #less than x runs -25
        bat_game['xPts'] = bat_game['xPts'] - 5*(rs.cdf(5)-rs.cdf(1))*(1-bf.cdf(30))
        #run rate bonus
        #bat_game['xPts'] = bat_game['xPts'] + 120*factor*bat_game['usage']*bat_game['runs/ball'] - (4/3)*(bat_game['usage']*factor*120)
        #SR bonuses
        bat_game['xPts'] = bat_game['xPts'] + 30*(1-SR.cdf(150))*(1-bf.cdf(20))
        bat_game['xPts'] = bat_game['xPts'] + 20*(SR.cdf(150)-SR.cdf(125))*(1-bf.cdf(20))
        bat_game['xPts'] = bat_game['xPts'] - 10*(SR.cdf(125)-SR.cdf(100))*(1-bf.cdf(20))
        bat_game['xPts'] = bat_game['xPts'] + 10*(SR.cdf(100)-SR.cdf(75))*(1-bf.cdf(20))
        bat_game['xPts'] = bat_game['xPts'] - 20*(SR.cdf(75)-SR.cdf(50))*(1-bf.cdf(20))
        bat_game['xPts'] = bat_game['xPts'] - 30*(SR.cdf(50)-SR.cdf(25))*(1-bf.cdf(20))
        bat_game['xPts'] = bat_game['xPts'] - 50*(SR.cdf(25))*(1-bf.cdf(20))
        #milestone bonuses
        bat_game['xPts'] = bat_game['xPts'] + 10*(rs.cdf(50)-rs.cdf(30))
        bat_game['xPts'] = bat_game['xPts'] + 20*(rs.cdf(75)-rs.cdf(50))
        bat_game['xPts'] = bat_game['xPts'] + 30*(rs.cdf(100)-rs.cdf(75))
        bat_game['xPts'] = bat_game['xPts'] + 50*(rs.cdf(150)-rs.cdf(100))
        bat_game['xPts'] = bat_game['xPts'] + 75*(1-rs.cdf(150))
        
        
    #bowler cant bowl more than 20% of the overs except in tests
    if(factor!=11.25):bowl_game['usage'] = np.minimum(bowl_game['usage'],0.2)
    bowl_game = bowl_game.drop(['pp usage'],axis=1)
    bowl_game = bowl_game.drop(['mid usage'],axis=1)
    bowl_game = bowl_game.drop(['setup usage'],axis=1)
    bowl_game = bowl_game.drop(['death usage'],axis=1)
    
    if(factor==5/6):
        #wicket is 35 pts
        bowl_game['xPts'] = 120*factor*bowl_game['usage']*bowl_game['wickets/ball']*35
        #dot balls pts
        bowl_game['xPts'] += 120*factor*bowl_game['usage']*bowl_game['dots/ball']*5
        #ECON bonuses
        bowl_game['xPts'] = bowl_game['xPts'] - 100*(1-ECON.cdf(8.5))*(1-bb.cdf(5))
        bowl_game['xPts'] = bowl_game['xPts'] - 80*(ECON.cdf(8.5)-ECON.cdf(7.5))*(1-bb.cdf(5))
        bowl_game['xPts'] = bowl_game['xPts'] - 40*(ECON.cdf(7.5)-ECON.cdf(7))*(1-bb.cdf(5))
        bowl_game['xPts'] = bowl_game['xPts'] + 60*(ECON.cdf(7)-ECON.cdf(6))*(1-bb.cdf(5))
        bowl_game['xPts'] = bowl_game['xPts'] + 80*(ECON.cdf(6)-ECON.cdf(5))*(1-bb.cdf(5))
        bowl_game['xPts'] = bowl_game['xPts'] + 100*(ECON.cdf(5)-ECON.cdf(4))*(1-bb.cdf(5))
        bowl_game['xPts'] = bowl_game['xPts'] + 120*(ECON.cdf(3))*(1-bb.cdf(5))
        #milestone bonuses
        #bowl_game['xPts'] = bowl_game['xPts'] + 30*(wkts.cdf(3)-wkts.cdf(2))
        bowl_game['xPts'] = bowl_game['xPts'] + 80*(wkts.cdf(4)-wkts.cdf(3))
        bowl_game['xPts'] = bowl_game['xPts'] + 120*(wkts.cdf(5)-wkts.cdf(4))
        bowl_game['xPts'] = bowl_game['xPts'] + 150*(wkts.cdf(6)-wkts.cdf(5))
        bowl_game['xPts'] = bowl_game['xPts'] + 200*(1-wkts.cdf(6))
        #maiden over
        bowl_game['xPts'] = bowl_game['xPts'] + 100*np.power(bowl_game['dots/ball'],5)*20*bowl_game['usage']*factor
        
    elif(factor == 1):
        #wicket is 35 pts
        bowl_game['xPts'] = 120*factor*bowl_game['usage']*bowl_game['wickets/ball']*35
        #dot balls pts
        bowl_game['xPts'] += 120*factor*bowl_game['usage']*bowl_game['dots/ball']*5
        #ECON bonuses
        bowl_game['xPts'] = bowl_game['xPts'] - 100*(1-ECON.cdf(9.5))*(1-bb.cdf(6))
        bowl_game['xPts'] = bowl_game['xPts'] - 80*(ECON.cdf(9.5)-ECON.cdf(8.5))*(1-bb.cdf(6))
        bowl_game['xPts'] = bowl_game['xPts'] - 40*(ECON.cdf(8.5)-ECON.cdf(8))*(1-bb.cdf(6))
        bowl_game['xPts'] = bowl_game['xPts'] + 60*(ECON.cdf(8)-ECON.cdf(7))*(1-bb.cdf(6))
        bowl_game['xPts'] = bowl_game['xPts'] + 80*(ECON.cdf(7)-ECON.cdf(6))*(1-bb.cdf(6))
        bowl_game['xPts'] = bowl_game['xPts'] + 100*(ECON.cdf(6)-ECON.cdf(4))*(1-bb.cdf(6))
        bowl_game['xPts'] = bowl_game['xPts'] + 120*(ECON.cdf(4))*(1-bb.cdf(6))
        #milestone bonuses
        #bowl_game['xPts'] = bowl_game['xPts'] + 30*(wkts.cdf(3)-wkts.cdf(2))
        bowl_game['xPts'] = bowl_game['xPts'] + 80*(wkts.cdf(4)-wkts.cdf(3))
        bowl_game['xPts'] = bowl_game['xPts'] + 120*(wkts.cdf(5)-wkts.cdf(4))
        bowl_game['xPts'] = bowl_game['xPts'] + 150*(wkts.cdf(6)-wkts.cdf(5))
        bowl_game['xPts'] = bowl_game['xPts'] + 200*(1-wkts.cdf(6))
        #maiden over
        bowl_game['xPts'] = bowl_game['xPts'] + 100*np.power(bowl_game['dots/ball'],6)*20*bowl_game['usage']*factor
        
    elif(factor == 2.5):
        #wicket is 30 pts
        bowl_game['xPts'] = 120*factor*bowl_game['usage']*bowl_game['wickets/ball']*35
        #dot balls pts
        #bowl_game['xPts'] += 120*factor*bowl_game['usage']*bowl_game['dots/ball']*5
        #ECON bonuses
        bowl_game['xPts'] = bowl_game['xPts'] - 40*(1-ECON.cdf(9))*(1-bb.cdf(18))
        bowl_game['xPts'] = bowl_game['xPts'] - 20*(ECON.cdf(9)-ECON.cdf(9))*(1-bb.cdf(18))
        bowl_game['xPts'] = bowl_game['xPts'] - 10*(ECON.cdf(7)-ECON.cdf(6))*(1-bb.cdf(18))
        bowl_game['xPts'] = bowl_game['xPts'] + 10*(ECON.cdf(6)-ECON.cdf(5))*(1-bb.cdf(18))
        bowl_game['xPts'] = bowl_game['xPts'] + 20*(ECON.cdf(5)-ECON.cdf(4))*(1-bb.cdf(18))
        bowl_game['xPts'] = bowl_game['xPts'] + 30*(ECON.cdf(4)-ECON.cdf(3))*(1-bb.cdf(18))
        bowl_game['xPts'] = bowl_game['xPts'] + 50*(ECON.cdf(3))*(1-bb.cdf(18))
        #milestone bonuses
        bowl_game['xPts'] = bowl_game['xPts'] + 10*(wkts.cdf(3)-wkts.cdf(2))
        bowl_game['xPts'] = bowl_game['xPts'] + 30*(wkts.cdf(4)-wkts.cdf(3))
        bowl_game['xPts'] = bowl_game['xPts'] + 50*(wkts.cdf(5)-wkts.cdf(4))
        bowl_game['xPts'] = bowl_game['xPts'] + 70*(wkts.cdf(6)-wkts.cdf(5))
        bowl_game['xPts'] = bowl_game['xPts'] + 100*(1-wkts.cdf(6))
        #maiden over
        bowl_game['xPts'] = bowl_game['xPts'] + 15*np.power(bowl_game['dots/ball'],6)*20*bowl_game['usage']*factor
    

    runs_t1,runs_t2 = game_results(bat_game,factor,t1,t2,s1,s2,c1,c2,league_avg,summary)
    global game_avg
    game_avg.append((runs_t1+runs_t2)/2)
    return (bat_game,bowl_game,summary,runs_t1,runs_t2)

def final_projections(bat_game,bowl_game,runs_t1,runs_t2):
    names = np.unique(np.concatenate([bat_game['batsman'].values,bowl_game['bowler'].values]))
    final = [["player","team","total","bat","bowl","bat usage","bowl usage","+/-"]]

    for x in names:
        try : p_bat = bat_game.loc[(bat_game['batsman']==x),'team'].values[0]  
        except IndexError: p_bat = "Free Agent"
        try : p_bowl = bowl_game.loc[(bowl_game['bowler']==x),'team'].values[0]
        except IndexError: p_bowl = "Free Agent"
        if(p_bat == "Free Agent"): p_team = p_bowl
        else : p_team = p_bat
        
        bat_game = bat_game.groupby(['batsman','season','team'], as_index=False).sum()
        bowl_game = bowl_game.groupby(['bowler','season','team'], as_index=False).sum()
        #partimers removed from bowling list, min 1 over in t20s, review this
        if(coversoff == 1 or cricdraft == 1): bowl_game.drop(bowl_game[bowl_game['usage'] < 0.025].index, inplace = True)
        else: bowl_game.drop(bowl_game[bowl_game['usage'] < 0.01].index, inplace = True)
        final.append([x,p_team,0,bat_game.loc[(bat_game['batsman']==x),'xPts'].mean(),bowl_game.loc[(bowl_game['bowler']==x),'xPts'].mean(),bat_game.loc[(bat_game['batsman']==x),'usage'].mean(),bowl_game.loc[(bowl_game['bowler']==x),'usage'].mean()])
        
    final = pd.DataFrame(final)
    final.columns = final.iloc[0];final = final.drop(0)
    final = final.fillna(0)
    if(coversoff == 1): final['total'] = final['bat'] + final['bowl'] + 25
    elif(ex22 == 1 or cricdraft == 1): final['total'] = final['bat'] + final['bowl']
    else: final['total'] = final['bat'] + final['bowl'] + 4
    final = final.sort_values(['total'], ascending=[False])
    
    for x in names:
        final.loc[final['player']==x,'bat +/-'] = 1.2*f*bat_game.loc[bat_game['batsman']==x,'usage'].sum()*(bat_game.loc[bat_game['batsman']==x,'SR'].sum()-bat_game.loc[bat_game['batsman']==x,'xSR'].sum())
        final.loc[final['player']==x,'bowl +/-'] = 20*f*bowl_game.loc[bowl_game['bowler']==x,'usage'].sum()*(bowl_game.loc[bowl_game['bowler']==x,'ECON'].sum()-bowl_game.loc[bowl_game['bowler']==x,'xECON'].sum())
    
    c=0
    while c<len(home):
        pm_t1_bat = final.loc[final['team']==home[c],'bat +/-'].sum()
        pm_t1_bowl = final.loc[final['team']==home[c],'bowl +/-'].sum()
        pm_t2_bat = final.loc[final['team']==opps[c],'bat +/-'].sum()
        pm_t2_bowl = final.loc[final['team']==opps[c],'bowl +/-'].sum()
        final.loc[final['team']==home[c],'bat +/-'] = final.loc[final['team']==home[c],'bat +/-'] + ((runs_t1[c]-game_avg[c]-pm_t1_bat) * final.loc[final['team']==home[c],'bat usage'])
        final.loc[final['team']==opps[c],'bat +/-'] = final.loc[final['team']==opps[c],'bat +/-'] + ((runs_t2[c]-game_avg[c]-pm_t2_bat) * final.loc[final['team']==opps[c],'bat usage'])
        #print(pm_t1_bowl,runs_t2[c]-game_avg[c],pm_t2_bowl,runs_t1[c]-game_avg[c])
        final.loc[final['team']==home[c],'bowl +/-'] = final.loc[final['team']==home[c],'bowl +/-'] + ((runs_t2[c]-game_avg[c]-pm_t1_bowl) * final.loc[final['team']==home[c],'bowl usage'])
        final.loc[final['team']==opps[c],'bowl +/-'] = final.loc[final['team']==opps[c],'bowl +/-'] + ((runs_t1[c]-game_avg[c]-pm_t2_bowl) * final.loc[final['team']==opps[c],'bowl usage'])
        c+=1
    final['+/-'] = (final['bat +/-']) - (final['bowl +/-'])    
    final['+/-'] = round(final['+/-'],2)
    #print(final.loc[final['team']==home[0],'bat +/-'].sum())
    #print(final.loc[final['team']==opps[0],'bat +/-'].sum())
    #print(final.loc[final['team']==home[0],'bowl +/-'].sum())
    #print(final.loc[final['team']==opps[0],'bowl +/-'].sum())
    final.drop(['bat +/-','bowl +/-'], axis=1, inplace=True)
    return final

def execute_projections():
    c = 0;
    for x in home:
        a = home[c]
        b = opps[c]
        v = venue[c]
        if (c==0):
            if(coversoff == 1): (bat_game,bowl_game,table,runs_t1,runs_t2) = coversoff_projection(a,b,input_file1,input_file,f,v)
            elif(cricdraft == 1): (bat_game,bowl_game,table,runs_t1,runs_t2) = cricdraft_projection(a,b,input_file1,input_file,f,v)
            elif(ex22 == 1): (bat_game,bowl_game,table,runs_t1,runs_t2) = ex22_projection(a,b,input_file1,input_file,f,v)
            else: (bat_game,bowl_game,table,runs_t1,runs_t2) = gw_projection(a,b,input_file1,input_file,f,v)
            runs_t1 = [runs_t1]
            runs_t2 = [runs_t2]
        else : 
            if(coversoff == 1): (bat,bowl,table,r_t1,r_t2) = coversoff_projection(a,b,input_file1,input_file,f,v)
            elif(cricdraft == 1): (bat,bowl,table,r_t1,r_t2) = cricdraft_projection(a,b,input_file1,input_file,f,v)
            elif(ex22 == 1): (bat,bowl,table,r_t1,r_t2) = ex22_projection(a,b,input_file1,input_file,f,v)
            else: (bat,bowl,table,r_t1,r_t2) = gw_projection(a,b,input_file1,input_file,f,v)
            bat_game = pd.concat([bat_game,bat])
            bowl_game = pd.concat([bowl_game,bowl])
            runs_t1.append(r_t1)
            runs_t2.append(r_t2)
        c = c + 1
    a_final = final_projections(bat_game,bowl_game,runs_t1,runs_t2)
    return a_final,table,bat_game,bowl_game

a_final,a_table,bat_game,bowl_game = execute_projections()
#a_final.rename(columns = {'bat usage':'balls faced', 'bowl usage':'overs bowled'}, inplace = True)
#a_final['balls faced'] = a_final['balls faced']*120*f
#a_final['overs bowled'] = a_final['overs bowled']*20*f
#a_final['overs bowled'] = (a_final['overs bowled'] - (a_final['overs bowled'] % 1)) + ((a_final['overs bowled'] % 1)*0.6)

#%% create a dsitribution of players
delta = 1; k=0
a_final['iters'] = pow(a_final['total'],3)
tot = a_final['iters'].sum()
a_final['iters'] = a_final['iters']/tot
a_final['iters'] = a_final['iters']*10*unique_combos
a_final['iters'] = round(np.minimum(a_final['iters'],unique_combos))
delta = a_final['iters'].sum() - 10*unique_combos
while(delta>0):
    a_final['iters'][len(a_final)-k] = a_final['iters'][len(a_final)-k] - 1
    delta = a_final['iters'].sum() - 10*unique_combos
    k+=1     

#%% generate n unique combos
def random_partition(seq, k):
    cnts = Counter(seq)
    # as long as there are enough items to "sample" take a random sample
    while len(cnts) >= k:
        sample = random.sample(list(cnts), k)
        cnts -= Counter(sample)
        yield sample

    # Fewer different items than the sample size, just return the unique
    # items until the Counter is empty
    while cnts:
        sample = list(cnts)
        cnts -= Counter(sample)
        yield sample

def distributer(a_projection,home,opps,unique_teams):
    team = [["1","2","3","4","5","6","7","8","9","10","11"]]; length = 0
    p_list = a_projection.reindex(a_projection.index.repeat(a_projection.iters))
    p_list = p_list['player'].tolist()
    #print(list(random_partition(p_list,10)))
    while(length != unique_teams):
        master_list = list(random_partition(p_list,10))
        length = len(master_list)
    for x in master_list:
        team.append(x)
    team = pd.DataFrame(team)
    team.columns = team.iloc[0];team = team.drop(0)
    team = team.T
    team = team.fillna("")
    return team

def randomizer(a_projection,home,opps,unique_teams):
    team = [["1","2","3","4","5","6","7","8","9","10","total"]]; i=0; j=0;
    players = a_projection.loc[(a_projection['team'] == home) | (a_projection['team'] == opps)]
    p = pow(players['total'],3).tolist()
    players = players['player'].tolist()
    sum_p = sum(p)
    p = [x/sum_p for x in p]
    
    while i<unique_teams:
        h=0; o=0;
        combo = np.random.choice(players, 10, p=p, replace=False)
        combo = combo.tolist()
        combo = sorted(combo)
        while j<10:
            t = a_projection.loc[a_projection['player'] == combo[j], 'team'].values[0]
            if(t==home): h+=1
            if(t==opps): o+=1
            j+=1
        if(h>10 or o>10): i=i-1
        else:
            cap = a_projection[a_projection.player.isin(combo)]
            pts = sum(cap['total'])
            p2 = pow(cap['total'],4).tolist()
            cap = cap['player'].tolist()
            sum_p2 = sum(p2)
            p2 = [x/sum_p2 for x in p2]
            y = np.random.choice(cap, 2, p=p2, replace=False)
            y = y.tolist()        
            #pts += a_projection.loc[(a_projection['player']==y[0]),'total'].sum() + (a_projection.loc[(a_projection['player']==y[1]),'total'].sum()/2)
            #combo += y + [pts]
            combo += [pts]
            team.append(combo)
            
        i +=1; j=0
        
    team = pd.DataFrame(team)
    team.columns = team.iloc[0];team = team.drop(0)
    team = team.T
    return team

a_final['count'] = 0
if(ex22 == 0 and coversoff == 0):
    c=0;combinations=[];
    while c<len(home):
        if(gw>0): co = randomizer(a_final,home[c],opps[c],unique_combos)
        else: co = randomizer(a_final,home[c],opps[c],unique_combos)
        combinations.append(co)
        for x in a_final['player'].values:
            a_final.loc[a_final['player']==x,'count'] = (co == x).values.sum()
        c+=1
a_final.drop('count', axis=1, inplace=True)

#%% dump projections to the desired file
date = now.strftime("%m-%d-%y")
def write_to_excel(date,output_dump,final):
    try:
        book = load_workbook(output_dump)
        with pd.ExcelWriter(output_dump, engine = 'openpyxl',mode='a', if_sheet_exists = 'replace') as writer:
            writer.book = book
            writer.sheets = {ws.title:ws for ws in book.worksheets}
            final.to_excel(writer, sheet_name=f'{date}',index=False)
    except FileNotFoundError:
        with pd.ExcelWriter(output_dump) as writer:        
            final.to_excel(writer, sheet_name=f'{date}',index=False)
            
def cricdraft_write(gw,output_dump,final):
    try:
        book = load_workbook(output_dump)
        with pd.ExcelWriter(output_dump, engine = 'openpyxl',mode='a', if_sheet_exists = 'replace') as writer:
            writer.book = book
            writer.sheets = {ws.title:ws for ws in book.worksheets}
            final.to_excel(writer, sheet_name = f'gw{gw}',index=False)
    except FileNotFoundError:
        with pd.ExcelWriter(output_dump) as writer:        
            final.to_excel(writer, sheet_name=f'gw{gw}', index=False)

if(write != 0 and gw == 0):
    write_to_excel(date,output_dump,a_final)
elif(coversoff == 0 and ex22 == 0 and cricdraft ==0 and write == 1):
    output_dump = f'{path}/outputs/{comp}_gw{gw}_{year}.xlsx'
    with pd.ExcelWriter(output_dump) as writer:        
        a_final.to_excel(writer, sheet_name='Points',index=False)
elif(coversoff == 1 and gw > 0 and write == 1):
    output_dump = f'{path}/outputs/{comp}_cricbattle_{year}.xlsx'
    with pd.ExcelWriter(output_dump) as writer:        
        a_final.to_excel(writer, sheet_name="Points", index=False)
elif(cricdraft == 1 and gw > 0 and write == 1):
    output_dump = f'{path}/outputs/{comp}_draft_{year}.xlsx'
    cricdraft_write(gw,output_dump,a_final)
